"""Ledgerly: a local, multi-company accounting ERP starter."""
from __future__ import annotations

import csv
import os
import sqlite3
from contextlib import closing
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from io import StringIO
from io import BytesIO

from flask import Flask, Response, flash, g, redirect, render_template, request, send_from_directory, session, url_for
from werkzeug.security import check_password_hash, generate_password_hash

ROOT = os.path.dirname(os.path.abspath(__file__))
DATABASE = os.path.join(ROOT, "ledger.db")
app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("ERP_SECRET_KEY", "change-this-before-production")
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(days=30)


@app.route("/service-worker.js")
def service_worker():
    """Serve the worker from the site root so it can support the whole ERP."""
    response = send_from_directory(os.path.join(ROOT, "static"), "service-worker.js", mimetype="application/javascript")
    response.headers["Service-Worker-Allowed"] = "/"
    response.headers["Cache-Control"] = "no-cache"
    return response


def normalise_date(value):
    """Accept DD-MM-YYYY (the ERP standard) and ISO dates; store ISO for safe sorting."""
    if isinstance(value, datetime): return value.date().isoformat()
    if isinstance(value, date): return value.isoformat()
    text = str(value or "").strip().split(" ")[0]
    for pattern in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%Y/%m/%d"):
        try: return datetime.strptime(text, pattern).date().isoformat()
        except ValueError: pass
    raise ValueError("Use date format DD-MM-YYYY")


def normalise_stock_timestamp(value):
    """Accept broker dates such as '06 Jul 2026, 11:02:38 AM' and keep FIFO time order."""
    if isinstance(value, datetime): return value.isoformat(timespec="seconds")
    if isinstance(value, date): return datetime.combine(value, datetime.min.time()).isoformat(timespec="seconds")
    text = str(value or "").strip().replace(",", "")
    try: return datetime.fromisoformat(text).isoformat(timespec="seconds")
    except ValueError: pass
    for pattern in ("%d %b %Y %I:%M:%S %p", "%d %B %Y %I:%M:%S %p", "%Y-%m-%d %H:%M:%S", "%d-%m-%Y %H:%M:%S"):
        try: return datetime.strptime(text, pattern).isoformat(timespec="seconds")
        except ValueError: pass
    return normalise_date(text) + "T00:00:00"


def display_date(value):
    try: return datetime.strptime(normalise_date(value), "%Y-%m-%d").strftime("%d-%m-%Y")
    except (ValueError, TypeError): return value or ""


@app.template_filter("datefmt")
def datefmt(value):
    return display_date(value)

DEFAULT_ACCOUNTS = [
    ("1000", "Bank account", "Asset", 1), ("1100", "Accounts receivable", "Asset", 0), ("1200", "Inventory", "Asset", 0), ("1500", "Equipment", "Asset", 0),
    ("2000", "Accounts payable", "Liability", 0), ("2100", "Taxes payable", "Liability", 0), ("3000", "Owner equity", "Equity", 0), ("3100", "Retained earnings", "Equity", 0),
    ("4000", "Sales revenue", "Income", 0), ("4100", "Service revenue", "Income", 0), ("5000", "Cost of sales", "Expense", 0), ("6000", "Operating expenses", "Expense", 0),
    ("6100", "Rent expense", "Expense", 0), ("6200", "Payroll expense", "Expense", 0),
]
DOCUMENT_TYPES = ("Journal", "Contra", "Sales", "Purchases", "Receipts", "Payments")
DEFAULT_SERIES = {"Journal": "JV-", "Contra": "CV-", "Sales": "SAL-", "Purchases": "PUR-", "Receipts": "RCV-", "Payments": "PAY-"}


def db():
    if "db" not in g:
        g.db = sqlite3.connect(DATABASE)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA foreign_keys = ON")
    return g.db


@app.teardown_appcontext
def close_db(_error=None):
    connection = g.pop("db", None)
    if connection:
        connection.close()


def money(value): return f"{Decimal(str(value or 0)):,.2f}"
app.jinja_env.filters["money"] = money


def init_db():
    schema = """
    CREATE TABLE IF NOT EXISTS companies (
      id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL UNIQUE, legal_name TEXT, tax_number TEXT,
      address TEXT, mobile TEXT, email TEXT, base_currency TEXT NOT NULL,
      document_number_mode TEXT NOT NULL DEFAULT 'automatic', document_prefix TEXT NOT NULL DEFAULT 'JV-', next_document_number INTEGER NOT NULL DEFAULT 1,
      financial_year_start TEXT NOT NULL, financial_year_end TEXT NOT NULL, auth_enabled INTEGER NOT NULL DEFAULT 0,
      created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
    );
    CREATE TABLE IF NOT EXISTS accounts (
      id INTEGER PRIMARY KEY AUTOINCREMENT, company_id INTEGER NOT NULL, code TEXT NOT NULL, name TEXT NOT NULL,
      category TEXT NOT NULL CHECK(category IN ('Asset','Liability','Equity','Income','Expense')),
      subgroup_id INTEGER, is_cash INTEGER NOT NULL DEFAULT 0, cash_type TEXT NOT NULL DEFAULT 'Bank', is_loan INTEGER NOT NULL DEFAULT 0, default_currency TEXT NOT NULL DEFAULT 'USD', active INTEGER NOT NULL DEFAULT 1, UNIQUE(company_id,code),
      FOREIGN KEY(company_id) REFERENCES companies(id) ON DELETE CASCADE
    );
    CREATE TABLE IF NOT EXISTS account_subgroups (
      id INTEGER PRIMARY KEY AUTOINCREMENT, company_id INTEGER NOT NULL, category TEXT NOT NULL CHECK(category IN ('Asset','Liability','Equity','Income','Expense')),
      name TEXT NOT NULL, UNIQUE(company_id,category,name), FOREIGN KEY(company_id) REFERENCES companies(id) ON DELETE CASCADE
    );
    CREATE TABLE IF NOT EXISTS accounting_tags (
      id INTEGER PRIMARY KEY AUTOINCREMENT, company_id INTEGER NOT NULL, name TEXT NOT NULL, analysis_category_id INTEGER, color TEXT NOT NULL DEFAULT '#007b9a', active INTEGER NOT NULL DEFAULT 1,
      UNIQUE(company_id,name), FOREIGN KEY(company_id) REFERENCES companies(id) ON DELETE CASCADE
    );
    CREATE TABLE IF NOT EXISTS analysis_categories (
      id INTEGER PRIMARY KEY AUTOINCREMENT, company_id INTEGER NOT NULL, name TEXT NOT NULL,
      UNIQUE(company_id,name), FOREIGN KEY(company_id) REFERENCES companies(id) ON DELETE CASCADE
    );
    CREATE TABLE IF NOT EXISTS account_analysis_categories (
      account_id INTEGER NOT NULL, category_id INTEGER NOT NULL, PRIMARY KEY(account_id,category_id),
      FOREIGN KEY(account_id) REFERENCES accounts(id) ON DELETE CASCADE, FOREIGN KEY(category_id) REFERENCES analysis_categories(id) ON DELETE CASCADE
    );
    CREATE TABLE IF NOT EXISTS account_tag_links (
      account_id INTEGER NOT NULL, tag_id INTEGER NOT NULL, PRIMARY KEY(account_id,tag_id),
      FOREIGN KEY(account_id) REFERENCES accounts(id) ON DELETE CASCADE, FOREIGN KEY(tag_id) REFERENCES accounting_tags(id) ON DELETE CASCADE
    );
    CREATE TABLE IF NOT EXISTS region_tags (
      id INTEGER PRIMARY KEY AUTOINCREMENT, company_id INTEGER NOT NULL, name TEXT NOT NULL, active INTEGER NOT NULL DEFAULT 1,
      UNIQUE(company_id,name), FOREIGN KEY(company_id) REFERENCES companies(id) ON DELETE CASCADE
    );
    CREATE TABLE IF NOT EXISTS company_settings (
      company_id INTEGER PRIMARY KEY, regions_enabled INTEGER NOT NULL DEFAULT 0,
      FOREIGN KEY(company_id) REFERENCES companies(id) ON DELETE CASCADE
    );
    CREATE TABLE IF NOT EXISTS journal_entries (
      id INTEGER PRIMARY KEY AUTOINCREMENT, company_id INTEGER NOT NULL, entry_date TEXT NOT NULL, document_type TEXT NOT NULL DEFAULT 'Journal', document_no TEXT, reference TEXT,
      memo TEXT, payment_mode TEXT, party TEXT, accounting_tag_id INTEGER, region_tag_id INTEGER, created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP, FOREIGN KEY(company_id) REFERENCES companies(id) ON DELETE CASCADE
    );
    CREATE TABLE IF NOT EXISTS journal_lines (
      id INTEGER PRIMARY KEY AUTOINCREMENT, entry_id INTEGER NOT NULL, account_id INTEGER NOT NULL, description TEXT,
      debit REAL NOT NULL DEFAULT 0, credit REAL NOT NULL DEFAULT 0, currency TEXT NOT NULL, fx_rate REAL NOT NULL DEFAULT 1,
      FOREIGN KEY(entry_id) REFERENCES journal_entries(id) ON DELETE CASCADE, FOREIGN KEY(account_id) REFERENCES accounts(id)
    );
    CREATE TABLE IF NOT EXISTS opening_balances (
      id INTEGER PRIMARY KEY AUTOINCREMENT, company_id INTEGER NOT NULL, account_id INTEGER NOT NULL, effective_date TEXT NOT NULL,
      debit REAL NOT NULL DEFAULT 0, credit REAL NOT NULL DEFAULT 0, currency TEXT NOT NULL DEFAULT 'INR', fx_rate REAL NOT NULL DEFAULT 1, UNIQUE(company_id,account_id,effective_date),
      FOREIGN KEY(company_id) REFERENCES companies(id) ON DELETE CASCADE, FOREIGN KEY(account_id) REFERENCES accounts(id) ON DELETE CASCADE
    );
    CREATE TABLE IF NOT EXISTS loan_profiles (
      account_id INTEGER PRIMARY KEY, company_id INTEGER NOT NULL, lender TEXT, principal REAL NOT NULL DEFAULT 0,
      annual_rate REAL NOT NULL DEFAULT 0, start_date TEXT, term_months INTEGER, notes TEXT,
      FOREIGN KEY(account_id) REFERENCES accounts(id) ON DELETE CASCADE
    );
    CREATE TABLE IF NOT EXISTS loan_repayments (
      id INTEGER PRIMARY KEY AUTOINCREMENT, account_id INTEGER NOT NULL, due_date TEXT NOT NULL,
      installment_no INTEGER, opening_principal REAL, installment_amount REAL,
      principal REAL NOT NULL DEFAULT 0, interest REAL NOT NULL DEFAULT 0, closing_principal REAL, paid INTEGER NOT NULL DEFAULT 0,
      FOREIGN KEY(account_id) REFERENCES accounts(id) ON DELETE CASCADE
    );
    CREATE TABLE IF NOT EXISTS bank_statement_lines (
      id INTEGER PRIMARY KEY AUTOINCREMENT, company_id INTEGER NOT NULL, account_id INTEGER NOT NULL,
      statement_date TEXT NOT NULL, description TEXT, reference TEXT, debit REAL NOT NULL DEFAULT 0, credit REAL NOT NULL DEFAULT 0,
      matched_journal_line_id INTEGER, created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
      FOREIGN KEY(company_id) REFERENCES companies(id) ON DELETE CASCADE, FOREIGN KEY(account_id) REFERENCES accounts(id),
      FOREIGN KEY(matched_journal_line_id) REFERENCES journal_lines(id)
    );
    CREATE TABLE IF NOT EXISTS bank_statement_openings (
      id INTEGER PRIMARY KEY AUTOINCREMENT, company_id INTEGER NOT NULL, account_id INTEGER NOT NULL,
      effective_date TEXT NOT NULL, debit REAL NOT NULL DEFAULT 0, credit REAL NOT NULL DEFAULT 0,
      UNIQUE(company_id,account_id),
      FOREIGN KEY(company_id) REFERENCES companies(id) ON DELETE CASCADE,
      FOREIGN KEY(account_id) REFERENCES accounts(id) ON DELETE CASCADE
    );
    CREATE TABLE IF NOT EXISTS currencies (
      id INTEGER PRIMARY KEY AUTOINCREMENT, company_id INTEGER NOT NULL, code TEXT NOT NULL, name TEXT NOT NULL,
      rate_to_base REAL NOT NULL DEFAULT 1, active INTEGER NOT NULL DEFAULT 1, UNIQUE(company_id,code),
      FOREIGN KEY(company_id) REFERENCES companies(id) ON DELETE CASCADE
    );
    CREATE TABLE IF NOT EXISTS document_series (
      id INTEGER PRIMARY KEY AUTOINCREMENT, company_id INTEGER NOT NULL, document_type TEXT NOT NULL,
      number_mode TEXT NOT NULL DEFAULT 'automatic', prefix TEXT NOT NULL, next_number INTEGER NOT NULL DEFAULT 1, allow_duplicates INTEGER NOT NULL DEFAULT 0,
      UNIQUE(company_id,document_type), FOREIGN KEY(company_id) REFERENCES companies(id) ON DELETE CASCADE
    );
    CREATE TABLE IF NOT EXISTS company_users (
      id INTEGER PRIMARY KEY AUTOINCREMENT, company_id INTEGER NOT NULL, username TEXT NOT NULL, contact_no TEXT, email TEXT, password_hash TEXT, role TEXT NOT NULL DEFAULT 'Accountant', active INTEGER NOT NULL DEFAULT 1,
      UNIQUE(company_id,username), FOREIGN KEY(company_id) REFERENCES companies(id) ON DELETE CASCADE
    );
    CREATE TABLE IF NOT EXISTS app_users (
      id INTEGER PRIMARY KEY AUTOINCREMENT, display_name TEXT NOT NULL, email TEXT NOT NULL UNIQUE COLLATE NOCASE,
      password_hash TEXT NOT NULL, active INTEGER NOT NULL DEFAULT 1, is_admin INTEGER NOT NULL DEFAULT 0,
      profession TEXT, industry TEXT, discovery_source TEXT, last_login_at TEXT, created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
    );
    CREATE TABLE IF NOT EXISTS company_access (
      app_user_id INTEGER NOT NULL, company_id INTEGER NOT NULL, role TEXT NOT NULL DEFAULT 'Owner',
      assigned_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP, PRIMARY KEY(app_user_id,company_id),
      FOREIGN KEY(app_user_id) REFERENCES app_users(id) ON DELETE CASCADE,
      FOREIGN KEY(company_id) REFERENCES companies(id) ON DELETE CASCADE
    );
    CREATE TABLE IF NOT EXISTS password_reset_requests (
      id INTEGER PRIMARY KEY AUTOINCREMENT, app_user_id INTEGER NOT NULL, requested_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
      status TEXT NOT NULL DEFAULT 'Pending', FOREIGN KEY(app_user_id) REFERENCES app_users(id) ON DELETE CASCADE
    );
    CREATE TABLE IF NOT EXISTS activity_log (
      id INTEGER PRIMARY KEY AUTOINCREMENT, company_id INTEGER NOT NULL, user_name TEXT, activity TEXT NOT NULL, details TEXT, created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
      FOREIGN KEY(company_id) REFERENCES companies(id) ON DELETE CASCADE
    );
    CREATE TABLE IF NOT EXISTS stock_transactions (
      id INTEGER PRIMARY KEY AUTOINCREMENT, company_id INTEGER NOT NULL, transaction_type TEXT NOT NULL CHECK(transaction_type IN ('BUY','SELL')),
      stock_name TEXT NOT NULL, symbol TEXT NOT NULL, quantity REAL NOT NULL CHECK(quantity > 0), transaction_date TEXT NOT NULL, transaction_timestamp TEXT,
      rate REAL NOT NULL CHECK(rate > 0), total_amount REAL NOT NULL CHECK(total_amount > 0), created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
      FOREIGN KEY(company_id) REFERENCES companies(id) ON DELETE CASCADE
    );
    CREATE TABLE IF NOT EXISTS stock_lots (
      id INTEGER PRIMARY KEY AUTOINCREMENT, company_id INTEGER NOT NULL, buy_transaction_id INTEGER NOT NULL,
      symbol TEXT NOT NULL, stock_name TEXT NOT NULL, purchase_date TEXT NOT NULL, quantity REAL NOT NULL,
      remaining_quantity REAL NOT NULL, rate REAL NOT NULL,
      FOREIGN KEY(company_id) REFERENCES companies(id) ON DELETE CASCADE,
      FOREIGN KEY(buy_transaction_id) REFERENCES stock_transactions(id) ON DELETE CASCADE
    );
    CREATE TABLE IF NOT EXISTS stock_holdings (
      id INTEGER PRIMARY KEY AUTOINCREMENT, company_id INTEGER NOT NULL, symbol TEXT NOT NULL, stock_name TEXT NOT NULL,
      remaining_quantity REAL NOT NULL DEFAULT 0, average_cost REAL NOT NULL DEFAULT 0, updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
      UNIQUE(company_id,symbol), FOREIGN KEY(company_id) REFERENCES companies(id) ON DELETE CASCADE
    );
    CREATE TABLE IF NOT EXISTS stock_realized_matches (
      id INTEGER PRIMARY KEY AUTOINCREMENT, company_id INTEGER NOT NULL, sale_transaction_id INTEGER NOT NULL, buy_lot_id INTEGER NOT NULL,
      symbol TEXT NOT NULL, sale_date TEXT NOT NULL, quantity REAL NOT NULL, sale_rate REAL NOT NULL, fifo_cost_rate REAL NOT NULL,
      realized_gain_loss REAL NOT NULL, holding_days INTEGER NOT NULL, tax_classification TEXT NOT NULL,
      FOREIGN KEY(company_id) REFERENCES companies(id) ON DELETE CASCADE,
      FOREIGN KEY(sale_transaction_id) REFERENCES stock_transactions(id) ON DELETE CASCADE,
      FOREIGN KEY(buy_lot_id) REFERENCES stock_lots(id) ON DELETE CASCADE
    );
    CREATE INDEX IF NOT EXISTS idx_stock_transactions_company_date ON stock_transactions(company_id,transaction_date,id);
    CREATE INDEX IF NOT EXISTS idx_stock_lots_company_symbol_date ON stock_lots(company_id,symbol,purchase_date,id);
    """
    with closing(sqlite3.connect(DATABASE)) as connection:
        old_account_columns = [row[1] for row in connection.execute("PRAGMA table_info(accounts)")]
        legacy = old_account_columns and "company_id" not in old_account_columns
        if legacy:
            connection.execute("PRAGMA foreign_keys = OFF")
            connection.execute("ALTER TABLE journal_lines RENAME TO legacy_journal_lines")
            connection.execute("ALTER TABLE journal_entries RENAME TO legacy_journal_entries")
            connection.execute("ALTER TABLE accounts RENAME TO legacy_accounts")
        connection.executescript(schema)
        columns = [row[1] for row in connection.execute("PRAGMA table_info(companies)")]
        for name in ("address", "mobile", "email"):
            if name not in columns:
                connection.execute(f"ALTER TABLE companies ADD COLUMN {name} TEXT")
        for name, definition in (("document_number_mode", "TEXT NOT NULL DEFAULT 'automatic'"), ("document_prefix", "TEXT NOT NULL DEFAULT 'JV-'"), ("next_document_number", "INTEGER NOT NULL DEFAULT 1"), ("edit_pin", "TEXT"), ("auth_enabled", "INTEGER NOT NULL DEFAULT 0")):
            if name not in columns:
                connection.execute(f"ALTER TABLE companies ADD COLUMN {name} {definition}")
        account_columns = [row[1] for row in connection.execute("PRAGMA table_info(accounts)")]
        if "cash_type" not in account_columns:
            connection.execute("ALTER TABLE accounts ADD COLUMN cash_type TEXT NOT NULL DEFAULT 'Bank'")
        if "default_currency" not in account_columns:
            connection.execute("ALTER TABLE accounts ADD COLUMN default_currency TEXT NOT NULL DEFAULT 'USD'")
        if "subgroup_id" not in account_columns:
            connection.execute("ALTER TABLE accounts ADD COLUMN subgroup_id INTEGER")
        if "is_loan" not in account_columns:
            connection.execute("ALTER TABLE accounts ADD COLUMN is_loan INTEGER NOT NULL DEFAULT 0")
        repayment_columns = [row[1] for row in connection.execute("PRAGMA table_info(loan_repayments)")]
        for name, definition in (("installment_no", "INTEGER"), ("opening_principal", "REAL"), ("installment_amount", "REAL"), ("closing_principal", "REAL")):
            if name not in repayment_columns:
                connection.execute(f"ALTER TABLE loan_repayments ADD COLUMN {name} {definition}")
        entry_columns = [row[1] for row in connection.execute("PRAGMA table_info(journal_entries)")]
        if "document_no" not in entry_columns:
            connection.execute("ALTER TABLE journal_entries ADD COLUMN document_no TEXT")
        if "document_type" not in entry_columns:
            connection.execute("ALTER TABLE journal_entries ADD COLUMN document_type TEXT NOT NULL DEFAULT 'Journal'")
        if "payment_mode" not in entry_columns:
            connection.execute("ALTER TABLE journal_entries ADD COLUMN payment_mode TEXT")
        if "party" not in entry_columns:
            connection.execute("ALTER TABLE journal_entries ADD COLUMN party TEXT")
        for name in ("accounting_tag_id", "region_tag_id"):
            if name not in entry_columns: connection.execute(f"ALTER TABLE journal_entries ADD COLUMN {name} INTEGER")
        tag_columns = [row[1] for row in connection.execute("PRAGMA table_info(accounting_tags)")]
        if "analysis_category_id" not in tag_columns: connection.execute("ALTER TABLE accounting_tags ADD COLUMN analysis_category_id INTEGER")
        connection.execute("INSERT OR IGNORE INTO company_settings(company_id) SELECT id FROM companies")
        opening_columns = [row[1] for row in connection.execute("PRAGMA table_info(opening_balances)")]
        if "currency" not in opening_columns:
            connection.execute("ALTER TABLE opening_balances ADD COLUMN currency TEXT NOT NULL DEFAULT 'INR'")
        if "fx_rate" not in opening_columns:
            connection.execute("ALTER TABLE opening_balances ADD COLUMN fx_rate REAL NOT NULL DEFAULT 1")
        if legacy:
            old_currency = connection.execute("SELECT value FROM settings WHERE key='base_currency'").fetchone()
            company = connection.execute("INSERT INTO companies(name,base_currency,financial_year_start,financial_year_end) VALUES(?,?,?,?)", ("Previous records", old_currency[0] if old_currency else "USD", "2000-01-01", "2099-12-31"))
            company_id = company.lastrowid
            connection.execute("INSERT INTO accounts(id,company_id,code,name,category,is_cash,active) SELECT id,?,code,name,category,is_cash,active FROM legacy_accounts", (company_id,))
            connection.execute("INSERT INTO journal_entries(id,company_id,entry_date,document_no,reference,memo,created_at) SELECT id,?,entry_date,'LEGACY-' || id,reference,memo,created_at FROM legacy_journal_entries", (company_id,))
            connection.execute("INSERT INTO journal_lines(id,entry_id,account_id,description,debit,credit,currency,fx_rate) SELECT id,entry_id,account_id,description,debit,credit,currency,fx_rate FROM legacy_journal_lines")
            connection.executescript("DROP TABLE legacy_journal_lines; DROP TABLE legacy_journal_entries; DROP TABLE legacy_accounts;")
            connection.execute("PRAGMA foreign_keys = ON")
        connection.execute("UPDATE journal_entries SET document_no='LEGACY-' || id WHERE document_no IS NULL OR trim(document_no)='' ")
        series_columns = [row[1] for row in connection.execute("PRAGMA table_info(document_series)")]
        if "allow_duplicates" not in series_columns:
            connection.execute("ALTER TABLE document_series ADD COLUMN allow_duplicates INTEGER NOT NULL DEFAULT 0")
        user_columns = [row[1] for row in connection.execute("PRAGMA table_info(company_users)")]
        for name, definition in (("contact_no", "TEXT"), ("email", "TEXT"), ("password_hash", "TEXT")):
            if name not in user_columns: connection.execute(f"ALTER TABLE company_users ADD COLUMN {name} {definition}")
        app_user_columns = [row[1] for row in connection.execute("PRAGMA table_info(app_users)")]
        if "is_admin" not in app_user_columns:
            connection.execute("ALTER TABLE app_users ADD COLUMN is_admin INTEGER NOT NULL DEFAULT 0")
        if "last_login_at" not in app_user_columns:
            connection.execute("ALTER TABLE app_users ADD COLUMN last_login_at TEXT")
        for name in ("profession", "industry", "discovery_source"):
            if name not in app_user_columns:
                connection.execute(f"ALTER TABLE app_users ADD COLUMN {name} TEXT")
        if not connection.execute("SELECT 1 FROM app_users WHERE is_admin=1 LIMIT 1").fetchone():
            connection.execute("UPDATE app_users SET is_admin=1 WHERE id=(SELECT id FROM app_users ORDER BY id LIMIT 1)")
        connection.execute("""INSERT OR IGNORE INTO company_access(app_user_id,company_id,role)
            SELECT au.id,cu.company_id,COALESCE(cu.role,'Accountant') FROM app_users au
            JOIN company_users cu ON lower(trim(cu.email))=lower(trim(au.email)) WHERE cu.active=1""")
        activity_columns = [row[1] for row in connection.execute("PRAGMA table_info(activity_log)")]
        if "user_name" not in activity_columns:
            connection.execute("ALTER TABLE activity_log ADD COLUMN user_name TEXT")
            connection.execute("UPDATE activity_log SET user_name=CASE WHEN instr(COALESCE(details,''),' · ')>0 THEN substr(details,1,instr(details,' · ')-1) WHEN trim(COALESCE(details,''))<>'' THEN details ELSE 'System' END")
            connection.execute("UPDATE activity_log SET details=CASE WHEN instr(COALESCE(details,''),' · ')>0 THEN substr(details,instr(details,' · ')+3) ELSE details END")
        connection.execute("DROP INDEX IF EXISTS idx_entry_document_no")
        for document_type in DOCUMENT_TYPES:
            connection.execute("INSERT OR IGNORE INTO document_series(company_id,document_type,number_mode,prefix,next_number) SELECT id,?,?,?,? FROM companies", (document_type, "automatic", DEFAULT_SERIES[document_type], 1))
        # INR is the standard base currency for every company and account.
        connection.execute("UPDATE companies SET base_currency='INR'")
        connection.execute("INSERT OR IGNORE INTO currencies(company_id,code,name,rate_to_base) SELECT id,'INR','Indian Rupee',1 FROM companies")
        connection.execute("UPDATE currencies SET name='Indian Rupee',rate_to_base=1 WHERE code='INR'")
        connection.execute("UPDATE currencies SET name=CASE code WHEN 'AED' THEN 'UAE Dirham' ELSE code || ' currency' END WHERE code <> 'INR' AND name LIKE '%base currency%'")
        connection.execute("UPDATE opening_balances SET currency=COALESCE((SELECT default_currency FROM accounts WHERE accounts.id=opening_balances.account_id),'INR') WHERE currency IS NULL OR currency='INR'")
        connection.execute("UPDATE opening_balances SET fx_rate=COALESCE((SELECT rate_to_base FROM currencies WHERE currencies.company_id=opening_balances.company_id AND currencies.code=opening_balances.currency),1) WHERE fx_rate IS NULL OR fx_rate=1")
        connection.execute("INSERT OR IGNORE INTO account_subgroups(company_id,category,name) SELECT id,'Asset','Bank Accounts' FROM companies")
        connection.execute("UPDATE accounts SET subgroup_id=(SELECT id FROM account_subgroups s WHERE s.company_id=accounts.company_id AND s.category='Asset' AND s.name='Bank Accounts') WHERE category='Asset' AND is_cash=1 AND cash_type='Bank' AND subgroup_id IS NULL")
        stock_columns = [row[1] for row in connection.execute("PRAGMA table_info(stock_transactions)")]
        if "transaction_timestamp" not in stock_columns:
            connection.execute("ALTER TABLE stock_transactions ADD COLUMN transaction_timestamp TEXT")
            connection.execute("UPDATE stock_transactions SET transaction_timestamp=transaction_date WHERE transaction_timestamp IS NULL")
        connection.commit()


def visible_companies():
    user = signed_in_user()
    if not user:
        return []
    ensure_user_company_access(user)
    return db().execute("""SELECT c.* FROM companies c
        JOIN company_access ca ON ca.company_id=c.id
        WHERE ca.app_user_id=? AND c.name <> 'Imported company' ORDER BY c.name""", (user["id"],)).fetchall()


def currency_list(company):
    if not company: return []
    db().execute("INSERT OR IGNORE INTO currencies(company_id,code,name,rate_to_base) VALUES(?,?,?,1)", (company["id"], company["base_currency"], f"{company['base_currency']} base currency")); db().commit()
    return db().execute("SELECT * FROM currencies WHERE company_id=? AND active=1 ORDER BY CASE WHEN code=? THEN 0 ELSE 1 END, code", (company["id"], company["base_currency"])).fetchall()


def active_company():
    company_id = session.get("company_id")
    if not company_id:
        return None
    return db().execute("SELECT * FROM companies WHERE id=? AND name <> 'Imported company'", (company_id,)).fetchone()


def signed_in_user():
    """Return the main Zedjer user stored in this browser session."""
    user_id = session.get("app_user_id")
    if not user_id:
        return None
    return db().execute("SELECT * FROM app_users WHERE id=? AND active=1", (user_id,)).fetchone()


def ensure_user_company_access(user):
    """Migrate matching legacy company users and give the first app owner their existing records."""
    connection = db()
    connection.execute("""INSERT OR IGNORE INTO company_access(app_user_id,company_id,role)
        SELECT ?,cu.company_id,COALESCE(cu.role,'Accountant') FROM company_users cu
        WHERE lower(trim(cu.email))=lower(trim(?)) AND cu.active=1""", (user["id"], user["email"]))
    has_access = connection.execute("SELECT 1 FROM company_access WHERE app_user_id=?", (user["id"],)).fetchone()
    user_count = connection.execute("SELECT COUNT(*) FROM app_users").fetchone()[0]
    if not has_access and user_count == 1:
        connection.execute("INSERT OR IGNORE INTO company_access(app_user_id,company_id,role) SELECT ?,id,'Owner' FROM companies WHERE name <> 'Imported company'", (user["id"],))
    connection.commit()


def user_company_access(user, company_id):
    ensure_user_company_access(user)
    return db().execute("SELECT * FROM company_access WHERE app_user_id=? AND company_id=?", (user["id"], company_id)).fetchone()


def company_required():
    user = signed_in_user()
    if not user:
        flash("Please sign in to continue.", "error")
        return None
    company = active_company()
    if company and not user_company_access(user, company["id"]):
        session.pop("company_id", None); session.pop("username", None); session.pop("user_role", None)
        flash("You do not have access to that company.", "error")
        return None
    return company


def document_series(company_id, document_type):
    return db().execute("SELECT * FROM document_series WHERE company_id=? AND document_type=?", (company_id, document_type)).fetchone()


def suggested_document_no(company, document_type="Journal"):
    series = document_series(company["id"], document_type)
    return f"{series['prefix'] or ''}{int(series['next_number'] or 1):05d}" if series else ""


def duplicate_document_number(company_id, document_type, document_no, exclude_entry_id=None):
    series = document_series(company_id, document_type)
    if series and series["allow_duplicates"]:
        return False
    sql, params = "SELECT 1 FROM journal_entries WHERE company_id=? AND document_type=? AND document_no=?", [company_id, document_type, document_no]
    if exclude_entry_id:
        sql += " AND id<>?"; params.append(exclude_entry_id)
    return db().execute(sql, params).fetchone() is not None


def audit(company_id, activity, details=""):
    actor = session.get("username", "System")
    db().execute("INSERT INTO activity_log(company_id,user_name,activity,details) VALUES(?,?,?,?)", (company_id, actor, activity, details))


def set_account_tags(company_id, account_id, tag_ids):
    valid = {row[0] for row in db().execute("SELECT id FROM accounting_tags WHERE company_id=?", (company_id,)).fetchall()}
    selected = {int(tag_id) for tag_id in tag_ids if str(tag_id).isdigit()} & valid
    db().execute("DELETE FROM account_tag_links WHERE account_id=?", (account_id,))
    db().executemany("INSERT INTO account_tag_links(account_id,tag_id) VALUES(?,?)", [(account_id, tag_id) for tag_id in selected])


def set_account_analysis_categories(company_id, account_id, category_ids):
    valid = {row[0] for row in db().execute("SELECT id FROM analysis_categories WHERE company_id=?", (company_id,)).fetchall()}
    selected = {int(category_id) for category_id in category_ids if str(category_id).isdigit()} & valid
    db().execute("DELETE FROM account_analysis_categories WHERE account_id=?", (account_id,))
    db().executemany("INSERT INTO account_analysis_categories(account_id,category_id) VALUES(?,?)", [(account_id, category_id) for category_id in selected])


def report_period(company):
    start = request.args.get("period_from") or request.args.get("from") or company["financial_year_start"]
    end = request.args.get("period_to") or request.args.get("to") or min(date.today().isoformat(), company["financial_year_end"])
    try:
        if datetime.strptime(start, "%Y-%m-%d") > datetime.strptime(end, "%Y-%m-%d"):
            raise ValueError
    except ValueError:
        start, end = company["financial_year_start"], min(date.today().isoformat(), company["financial_year_end"])
    return start, end


def report_currency(company):
    code = request.args.get("currency", "BASE").upper().strip()
    target = company["base_currency"] if code in ("", "BASE", company["base_currency"]) else code
    try:
        rate = Decimal(request.args.get("report_rate", "1"))
        if rate <= 0: raise InvalidOperation
    except (InvalidOperation, ValueError):
        rate = Decimal("1")
    return target, rate


def report_factor(company):
    target, rate = report_currency(company)
    return target, (Decimal("1") if target == company["base_currency"] else rate)


def export_report(company, title, headers, rows, currency_code):
    """Export an already prepared report table to a minimal branded PDF or Excel file."""
    export_type = request.args.get("export", "").lower()
    if export_type not in ("pdf", "excel"): return None
    metadata = [company["name"].upper(), company["address"] or "", f"Financial year: {display_date(company['financial_year_start'])} to {display_date(company['financial_year_end'])}", f"Currency: {currency_code}"]
    if export_type == "excel":
        try:
            from openpyxl import Workbook
        except ModuleNotFoundError:
            return Response("<h2>Excel export setup needed</h2><p>Please run: <code>python -m pip install -r requirements.txt</code>, then restart Ledgerly.</p>", status=503, mimetype="text/html")
        from openpyxl.styles import Font, PatternFill
        book = Workbook(); sheet = book.active; sheet.title = title[:31]
        for value in metadata: sheet.append([value])
        sheet.append([]); sheet.append(headers)
        for cell in sheet[sheet.max_row]: cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor="007B9A")
        for row in rows: sheet.append(row)
        for column in sheet.columns: sheet.column_dimensions[column[0].column_letter].width = min(max(len(str(c.value or "")) for c in column) + 2, 35)
        if title == "Chart of Accounts":
            for letter, width in zip(("A", "B", "C", "D", "E"), (16, 14, 42, 16, 16)):
                sheet.column_dimensions[letter].width = width
            for row in range(1, sheet.max_row + 1): sheet.row_dimensions[row].height = 20
            for row_number in range(7, sheet.max_row + 1):
                label = str(sheet.cell(row_number, 3).value or "")
                if label.startswith("TOTAL") or label == "FINAL TOTAL":
                    for cell in sheet[row_number]: cell.font = Font(bold=True, color="075E74"); cell.fill = PatternFill("solid", fgColor="EAF7FA")
                elif sheet.cell(row_number, 1).value:
                    for cell in sheet[row_number]: cell.fill = PatternFill("solid", fgColor="F7FBFC")
        output = BytesIO(); book.save(output)
        return Response(output.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={title.lower().replace(' ','-')}.xlsx"})
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    except ModuleNotFoundError:
        return Response("<h2>PDF export setup needed</h2><p>Please run: <code>python -m pip install -r requirements.txt</code>, then restart Ledgerly.</p>", status=503, mimetype="text/html")
    output = BytesIO(); doc = SimpleDocTemplate(output, pagesize=landscape(A4), rightMargin=24, leftMargin=24, topMargin=22, bottomMargin=22)
    styles = getSampleStyleSheet(); styles["Title"].textColor = colors.HexColor("#007B9A"); styles["Title"].fontSize = 20; styles["Heading2"].textColor = colors.HexColor("#075E74")
    story = [Paragraph(company["name"].upper(), styles["Title"]), Paragraph(company["address"] or "", styles["Normal"]), Paragraph(title, styles["Heading2"]), Paragraph(metadata[2] + " · " + metadata[3], styles["Normal"]), Spacer(1, 10)]
    table_data = [headers] + rows
    account_layout = title == "Chart of Accounts"
    table = Table(table_data, repeatRows=1, colWidths=[75, 65, 300, 85, 85] if account_layout else None, rowHeights=[20] * len(table_data) if account_layout else None)
    table.setStyle(TableStyle([("BACKGROUND", (0,0), (-1,0), colors.HexColor("#007B9A")), ("TEXTCOLOR", (0,0), (-1,0), colors.white), ("GRID", (0,0), (-1,-1), .25, colors.HexColor("#CFD8DC")), ("BACKGROUND", (0,1), (-1,-1), colors.white), ("FONTSIZE", (0,0), (-1,-1), 7), ("BOTTOMPADDING", (0,0), (-1,0), 7)]))
    if account_layout:
        account_style = [("BACKGROUND", (0, row_index), (-1, row_index), colors.HexColor("#EAF7FA")) if str(row[2] if len(row) > 2 else "").startswith("TOTAL") or str(row[2] if len(row) > 2 else "") == "FINAL TOTAL" else ("BACKGROUND", (0, row_index), (-1, row_index), colors.HexColor("#F7FBFC")) for row_index, row in enumerate(rows, start=1) if str(row[0] if row else "") or str(row[2] if len(row) > 2 else "").startswith("TOTAL")]
        account_style += [("FONTNAME", (0, row_index), (-1, row_index), "Helvetica-Bold") for row_index, row in enumerate(rows, start=1) if str(row[2] if len(row) > 2 else "").startswith("TOTAL") or str(row[2] if len(row) > 2 else "") == "FINAL TOTAL"]
        table.setStyle(TableStyle(account_style))
    story.append(table); doc.build(story)
    return Response(output.getvalue(), mimetype="application/pdf", headers={"Content-Disposition": f"attachment; filename={title.lower().replace(' ','-')}.pdf"})


def opening_adjustments(company_id, until_date):
    rows = db().execute("SELECT account_id,COALESCE(SUM((debit-credit)*fx_rate),0) amount FROM opening_balances WHERE company_id=? AND effective_date<=? GROUP BY account_id", (company_id, until_date)).fetchall()
    return {row["account_id"]: row["amount"] for row in rows}


def balances(company_id, as_of=None):
    as_of = as_of or "9999-12-31"
    rows = db().execute("""
      SELECT a.id,a.code,a.name,a.category,a.subgroup_id,a.is_cash,COALESCE(SUM(CASE WHEN e.id IS NOT NULL THEN (l.debit-l.credit)*l.fx_rate ELSE 0 END),0) signed_balance
      FROM accounts a LEFT JOIN journal_lines l ON l.account_id=a.id
      LEFT JOIN journal_entries e ON e.id=l.entry_id AND e.company_id=? AND e.entry_date<=?
      WHERE a.company_id=? GROUP BY a.id ORDER BY a.code
    """, (company_id, as_of, company_id)).fetchall()
    opening = opening_adjustments(company_id, as_of)
    output = []
    for row in rows:
        signed = row["signed_balance"] + opening.get(row["id"], 0)
        output.append({**dict(row), "signed_balance": signed, "balance": signed if row["category"] in ("Asset", "Expense") else -signed})
    return output


def account_balance_in_currency(company_id, account_id, currency, current_rate):
    """Return an account balance in its own displayed currency, preserving entered rates."""
    lines = db().execute("SELECT debit,credit,currency,fx_rate FROM journal_lines l JOIN journal_entries e ON e.id=l.entry_id WHERE e.company_id=? AND l.account_id=?", (company_id, account_id)).fetchall()
    openings = db().execute("SELECT debit,credit,currency,fx_rate FROM opening_balances WHERE company_id=? AND account_id=?", (company_id, account_id)).fetchall()
    total = 0
    for row in [*lines, *openings]:
        amount = row["debit"] - row["credit"]
        total += amount if row["currency"] == currency else amount * row["fx_rate"] / current_rate
    return total


def movements(company_id, start, end):
    rows = db().execute("""
      SELECT a.id,COALESCE(SUM(l.debit*l.fx_rate),0) debit,COALESCE(SUM(l.credit*l.fx_rate),0) credit
      FROM accounts a LEFT JOIN journal_lines l ON l.account_id=a.id
      LEFT JOIN journal_entries e ON e.id=l.entry_id AND e.company_id=? AND e.entry_date BETWEEN ? AND ?
      WHERE a.company_id=? GROUP BY a.id
    """, (company_id, start, end, company_id)).fetchall()
    return {row["id"]: (row["debit"], row["credit"]) for row in rows}


@app.context_processor
def global_values():
    company = active_company()
    return {"active_company": company, "companies": visible_companies(), "currencies": currency_list(company), "base_currency": company["base_currency"] if company else "", "today": date.today().isoformat(), "today_iso": date.today().isoformat()}


@app.route("/")
def dashboard():
    return render_template("opening.html")


@app.route("/signup", methods=["GET", "POST"])
def sign_up():
    if signed_in_user():
        return redirect(url_for("companies_dashboard"))
    if request.method == "POST":
        name = request.form.get("display_name", "").strip()
        email = request.form.get("email", "").strip().casefold()
        profession = request.form.get("profession", "").strip()
        industry = request.form.get("industry", "").strip()
        discovery_source = request.form.get("discovery_source", "").strip()
        password = request.form.get("password", "")
        confirm_password = request.form.get("confirm_password", "")
        try:
            if not name or not profession or not industry or not discovery_source or "@" not in email or "." not in email.rsplit("@", 1)[-1] or len(password) < 8 or password != confirm_password:
                raise ValueError
            is_first_user = db().execute("SELECT COUNT(*) FROM app_users").fetchone()[0] == 0
            cursor = db().execute(
                "INSERT INTO app_users(display_name,email,password_hash,is_admin,profession,industry,discovery_source,last_login_at) VALUES(?,?,?,?,?,?,?,CURRENT_TIMESTAMP)",
                (name, email, generate_password_hash(password), int(is_first_user), profession, industry, discovery_source),
            )
            db().commit()
            session.clear()
            session.permanent = True
            session["app_user_id"] = cursor.lastrowid
            ensure_user_company_access(signed_in_user())
            flash("Your Zedjer account is ready.", "success")
            return redirect(url_for("companies_dashboard"))
        except sqlite3.IntegrityError:
            flash("An account already exists with this email address. Please sign in.", "error")
        except ValueError:
            flash("Complete all setup details, use a valid email, and choose a password of at least 8 characters. Passwords must match.", "error")
    return render_template("signup.html")


@app.route("/signin", methods=["GET", "POST"])
def sign_in():
    if signed_in_user():
        return redirect(url_for("companies_dashboard"))
    if request.method == "POST":
        email = request.form.get("email", "").strip().casefold()
        password = request.form.get("password", "")
        user = db().execute("SELECT * FROM app_users WHERE email=? COLLATE NOCASE AND active=1", (email,)).fetchone()
        if user and check_password_hash(user["password_hash"], password):
            db().execute("UPDATE app_users SET last_login_at=CURRENT_TIMESTAMP WHERE id=?", (user["id"],))
            db().commit()
            session.clear()
            session.permanent = True
            session["app_user_id"] = user["id"]
            flash("Welcome back, " + user["display_name"] + ".", "success")
            return redirect(url_for("companies_dashboard"))
        if not user:
            flash("User not registered. Please create an account first.", "error")
        else:
            flash("Incorrect password. Please try again.", "error")
    return render_template("signin.html")


@app.route("/admin/signin", methods=["GET", "POST"])
def admin_sign_in():
    if signed_in_user() and signed_in_user()["is_admin"]:
        return redirect(url_for("admin_users"))
    if request.method == "POST":
        email = request.form.get("email", "").strip().casefold()
        password = request.form.get("password", "")
        user = db().execute("SELECT * FROM app_users WHERE email=? COLLATE NOCASE AND active=1 AND is_admin=1", (email,)).fetchone()
        if user and check_password_hash(user["password_hash"], password):
            db().execute("UPDATE app_users SET last_login_at=CURRENT_TIMESTAMP WHERE id=?", (user["id"],)); db().commit()
            session.clear(); session.permanent = True; session["app_user_id"] = user["id"]
            return redirect(url_for("admin_users"))
        flash("Administrator email or password is incorrect.", "error")
    return render_template("admin_signin.html")


@app.route("/forgot-password", methods=["GET", "POST"])
def forgot_password():
    if request.method == "POST":
        email = request.form.get("email", "").strip().casefold()
        user = db().execute("SELECT * FROM app_users WHERE email=? COLLATE NOCASE AND active=1", (email,)).fetchone()
        if not user:
            flash("User not registered. Please create an account first.", "error")
        else:
            pending = db().execute("SELECT 1 FROM password_reset_requests WHERE app_user_id=? AND status='Pending'", (user["id"],)).fetchone()
            if not pending:
                db().execute("INSERT INTO password_reset_requests(app_user_id) VALUES(?)", (user["id"],)); db().commit()
            flash("Your password reset request has been sent to the ERP administrator.", "success")
            return redirect(url_for("sign_in"))
    return render_template("forgot_password.html")


@app.before_request
def require_main_sign_in():
    """Keep every business workspace behind the main email/password sign-in."""
    public_endpoints = {"dashboard", "sign_in", "sign_up", "admin_sign_in", "forgot_password", "service_worker", "static", "user_logout"}
    if request.endpoint in public_endpoints or request.path.startswith("/static/"):
        return None
    if not signed_in_user():
        return redirect(url_for("sign_in"))
    return None


@app.route("/companies")
def companies_dashboard():
    return render_template("companies.html", show_form=False)


@app.route("/companies/new", methods=["GET", "POST"])
def company_new():
    if request.method == "POST":
        values = {key: request.form.get(key, "").strip() for key in ("name", "legal_name", "tax_number", "address", "mobile", "email", "financial_year_start", "financial_year_end")}
        currency = "INR"
        try:
            values["financial_year_start"], values["financial_year_end"] = normalise_date(values["financial_year_start"]), normalise_date(values["financial_year_end"])
            if not values["name"] or len(currency) != 3 or not currency.isalpha() or values["financial_year_start"] >= values["financial_year_end"]:
                raise ValueError
            connection = db()
            company = connection.execute("INSERT INTO companies(name,legal_name,tax_number,address,mobile,email,base_currency,financial_year_start,financial_year_end) VALUES(?,?,?,?,?,?,?,?,?)", (values["name"], values["legal_name"], values["tax_number"], values["address"], values["mobile"], values["email"], currency, values["financial_year_start"], values["financial_year_end"]))
            connection.executemany("INSERT INTO accounts(company_id,code,name,category,is_cash) VALUES(?,?,?,?,?)", [(company.lastrowid, *account) for account in DEFAULT_ACCOUNTS])
            connection.execute("INSERT INTO currencies(company_id,code,name,rate_to_base) VALUES(?,?,?,1)", (company.lastrowid, currency, f"{currency} base currency"))
            connection.executemany("INSERT INTO document_series(company_id,document_type,number_mode,prefix,next_number) VALUES(?,?,?,?,?)", [(company.lastrowid, kind, "automatic", DEFAULT_SERIES[kind], 1) for kind in DOCUMENT_TYPES])
            app_user = signed_in_user()
            if app_user:
                connection.execute("INSERT OR IGNORE INTO company_users(company_id,username,email,role,active) VALUES(?,?,?,?,1)", (company.lastrowid, app_user["display_name"], app_user["email"], "Owner"))
                connection.execute("INSERT OR IGNORE INTO company_access(app_user_id,company_id,role) VALUES(?,?,?)", (app_user["id"], company.lastrowid, "Owner"))
            connection.commit(); session["company_id"] = company.lastrowid
            flash("Company created. You can now record its transactions.", "success")
            return redirect(url_for("companies_dashboard"))
        except (ValueError, sqlite3.IntegrityError):
            flash("Enter a unique name, valid currency, and valid financial-year dates.", "error")
    return render_template("companies.html", show_form=True)


@app.route("/companies/<int:company_id>/edit", methods=["GET", "POST"])
def edit_company(company_id):
    company = db().execute("SELECT * FROM companies WHERE id=? AND name <> 'Imported company'", (company_id,)).fetchone()
    if not company: return redirect(url_for("companies_dashboard"))
    if request.method == "POST":
        try:
            name = request.form.get("name", "").strip(); currency = request.form.get("base_currency", "").upper().strip(); start = request.form.get("financial_year_start", ""); end = request.form.get("financial_year_end", "")
            pin = request.form.get("edit_pin", "").strip()
            start, end = normalise_date(start), normalise_date(end)
            if not name or len(currency) != 3 or not currency.isalpha() or start >= end or (pin and (len(pin) != 4 or not pin.isdigit())): raise ValueError
            db().execute("UPDATE companies SET name=?,legal_name=?,tax_number=?,address=?,mobile=?,email=?,base_currency=?,financial_year_start=?,financial_year_end=?,edit_pin=COALESCE(?,edit_pin) WHERE id=?", (name, request.form.get("legal_name", "").strip(), request.form.get("tax_number", "").strip(), request.form.get("address", "").strip(), request.form.get("mobile", "").strip(), request.form.get("email", "").strip(), currency, start, end, pin or None, company_id)); db().commit(); flash("Company details updated.", "success")
            return redirect(url_for("companies_dashboard"))
        except (ValueError, sqlite3.IntegrityError): flash("Use a unique company name and valid financial-year dates.", "error")
    return render_template("edit_company.html", company=company)


@app.route("/companies/<int:company_id>/document-numbering", methods=["GET", "POST"])
def document_numbering(company_id):
    company = db().execute("SELECT * FROM companies WHERE id=? AND name <> 'Imported company'", (company_id,)).fetchone()
    if not company: return redirect(url_for("companies_dashboard"))
    if request.method == "POST":
        try:
            connection = db()
            for document_type in DOCUMENT_TYPES:
                mode = request.form.get(f"mode_{document_type}", "automatic")
                prefix = request.form.get(f"prefix_{document_type}", "").strip()
                next_number = int(request.form.get(f"next_{document_type}", "0"))
                if mode not in ("automatic", "manual") or next_number < 1: raise ValueError
                connection.execute("UPDATE document_series SET number_mode=?,prefix=?,next_number=?,allow_duplicates=? WHERE company_id=? AND document_type=?", (mode, prefix, next_number, int(bool(request.form.get(f"duplicates_{document_type}"))), company_id, document_type))
            connection.commit(); flash("Numbering series saved.", "success")
            return redirect(url_for("master"))
        except ValueError: flash("Choose a valid numbering option and next number.", "error")
    return render_template("document_numbering.html", company=company, series={row["document_type"]: row for row in db().execute("SELECT * FROM document_series WHERE company_id=?", (company_id,))})


@app.route("/document-series/<document_type>")
def document_series_info(document_type):
    company = company_required()
    if not company or document_type not in DOCUMENT_TYPES:
        return {"error": "Not found"}, 404
    series = document_series(company["id"], document_type)
    return {"mode": series["number_mode"], "suggested": suggested_document_no(company, document_type)}


@app.route("/master")
def master():
    company = company_required()
    if not company: return redirect(url_for("companies_dashboard"))
    return render_template("master.html")


@app.route("/accounts-workspace")
def accounts_workspace():
    company = company_required()
    if not company: return redirect(url_for("companies_dashboard"))
    return render_template("accounts_workspace.html")


def rebuild_stock_portfolio(connection, company_id):
    """Rebuild lots, FIFO matches and grouped holdings from the source transactions."""
    connection.execute("DELETE FROM stock_realized_matches WHERE company_id=?", (company_id,))
    connection.execute("DELETE FROM stock_lots WHERE company_id=?", (company_id,))
    connection.execute("DELETE FROM stock_holdings WHERE company_id=?", (company_id,))
    transactions = connection.execute("SELECT * FROM stock_transactions WHERE company_id=? ORDER BY COALESCE(transaction_timestamp,transaction_date),id", (company_id,)).fetchall()
    lots_by_symbol = {}
    for transaction in transactions:
        symbol = transaction["symbol"]
        quantity = Decimal(str(transaction["quantity"]))
        rate = Decimal(str(transaction["rate"]))
        if transaction["transaction_type"] == "BUY":
            cursor = connection.execute("INSERT INTO stock_lots(company_id,buy_transaction_id,symbol,stock_name,purchase_date,quantity,remaining_quantity,rate) VALUES(?,?,?,?,?,?,?,?)", (company_id, transaction["id"], symbol, transaction["stock_name"], transaction["transaction_date"], float(quantity), float(quantity), float(rate)))
            lots_by_symbol.setdefault(symbol, []).append({"id": cursor.lastrowid, "date": transaction["transaction_date"], "remaining": quantity, "rate": rate, "name": transaction["stock_name"]})
            continue
        remaining_to_sell = quantity
        available_lots = lots_by_symbol.get(symbol, [])
        if sum((lot["remaining"] for lot in available_lots), Decimal("0")) + Decimal("0.0000001") < remaining_to_sell:
            raise ValueError(f"Cannot sell {symbol}: quantity exceeds the available holding.")
        sale_date = date.fromisoformat(transaction["transaction_date"])
        for lot in available_lots:
            if remaining_to_sell <= 0: break
            matched = min(lot["remaining"], remaining_to_sell)
            if matched <= 0: continue
            lot["remaining"] -= matched
            remaining_to_sell -= matched
            holding_days = (sale_date - date.fromisoformat(lot["date"])).days
            classification = "Long-Term" if holding_days > 365 else "Short-Term"
            gain_loss = (rate - lot["rate"]) * matched
            connection.execute("UPDATE stock_lots SET remaining_quantity=? WHERE id=?", (float(lot["remaining"]), lot["id"]))
            connection.execute("INSERT INTO stock_realized_matches(company_id,sale_transaction_id,buy_lot_id,symbol,sale_date,quantity,sale_rate,fifo_cost_rate,realized_gain_loss,holding_days,tax_classification) VALUES(?,?,?,?,?,?,?,?,?,?,?)", (company_id, transaction["id"], lot["id"], symbol, transaction["transaction_date"], float(matched), float(rate), float(lot["rate"]), float(gain_loss), holding_days, classification))
    for symbol, lots in lots_by_symbol.items():
        open_lots = [lot for lot in lots if lot["remaining"] > 0]
        quantity = sum((lot["remaining"] for lot in open_lots), Decimal("0"))
        if quantity <= 0: continue
        total_cost = sum((lot["remaining"] * lot["rate"] for lot in open_lots), Decimal("0"))
        connection.execute("INSERT INTO stock_holdings(company_id,symbol,stock_name,remaining_quantity,average_cost) VALUES(?,?,?,?,?)", (company_id, symbol, open_lots[0]["name"], float(quantity), float(total_cost / quantity)))


@app.route("/investment-analysis", methods=["GET", "POST"])
def investment_analysis():
    company = company_required()
    if not company: return redirect(url_for("companies_dashboard"))
    if request.method == "POST":
        if request.form.get("form_kind") == "manual":
            try:
                stock_name = request.form.get("stock_name", "").strip()
                symbol = "".join(character for character in stock_name.upper() if character.isalnum())[:24]
                transaction_type = request.form.get("transaction_type", "").upper().strip()
                quantity, rate, total = Decimal(request.form.get("quantity") or 0), Decimal(request.form.get("rate") or 0), Decimal(request.form.get("total_amount") or 0)
                timestamp = normalise_stock_timestamp(request.form.get("transaction_timestamp", ""))
                if transaction_type not in ("BUY", "SELL") or not stock_name or not symbol or quantity <= 0 or rate <= 0 or total <= 0 or abs((quantity * rate) - total) > Decimal("0.02"): raise ValueError("Enter a valid BUY or SELL transaction and ensure Total value equals Quantity × Rate.")
                connection = db(); connection.execute("INSERT INTO stock_transactions(company_id,transaction_type,stock_name,symbol,quantity,transaction_date,transaction_timestamp,rate,total_amount) VALUES(?,?,?,?,?,?,?,?,?)", (company["id"], transaction_type, stock_name, symbol, float(quantity), timestamp[:10], timestamp, float(rate), float(total)))
                rebuild_stock_portfolio(connection, company["id"]); audit(company["id"], "Stock investment added", f"{transaction_type} {stock_name}"); connection.commit(); flash("Stock investment saved and FIFO holdings updated.", "success")
            except (ValueError, InvalidOperation) as error:
                db().rollback(); flash("Unable to save investment. " + str(error), "error")
            except sqlite3.Error:
                db().rollback(); flash("Unable to save investment. Please try again.", "error")
            return redirect(url_for("investment_analysis"))
        upload = request.files.get("file")
        try:
            if not upload or not upload.filename.lower().endswith((".xlsx", ".xls")): raise ValueError("Choose an Excel file.")
            try: import pandas as pd
            except ModuleNotFoundError:
                # The normal engine uses pandas; this compatible fallback keeps imports
                # working in installations that only have the existing openpyxl package.
                try:
                    from openpyxl import load_workbook
                    required = ["Transaction_Type", "Stock_Name", "Symbol", "Quantity", "Purchase_Date", "Rate", "Total_Amount"]
                    broker_columns = ["Name", "Date & time", "Side", "Qty", "Traded price", "Total value", "Segment"]
                    sheet = load_workbook(upload, data_only=True).active
                    headers = [str(cell.value or "").strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
                    broker_format = all(column in headers for column in broker_columns)
                    if not broker_format and any(column not in headers for column in required): raise ValueError("Use the downloadable broker template or the standard investment template.")
                    indexes = {column: headers.index(column) for column in (broker_columns if broker_format else required)}; rows = []
                    for values in sheet.iter_rows(min_row=2, values_only=True):
                        if not any(value is not None and str(value).strip() for value in values): continue
                        transaction_type = str(values[indexes["Side" if broker_format else "Transaction_Type"]] or "").strip().upper(); stock_name = str(values[indexes["Name" if broker_format else "Stock_Name"]] or "").strip()
                        symbol = ("".join(character for character in stock_name.upper() if character.isalnum())[:24] if broker_format else str(values[indexes["Symbol"]] or "").strip().upper())
                        quantity, rate, total = (Decimal(str(values[indexes["Qty" if broker_format else "Quantity"]] or 0).replace(",", "")), Decimal(str(values[indexes["Traded price" if broker_format else "Rate"]] or 0).replace(",", "")), Decimal(str(values[indexes["Total value" if broker_format else "Total_Amount"]] or 0).replace(",", "")))
                        timestamp = normalise_stock_timestamp(values[indexes["Date & time" if broker_format else "Purchase_Date"]]); transaction_date = timestamp[:10]
                        if transaction_type not in ("BUY", "SELL") or not stock_name or not symbol or quantity <= 0 or rate <= 0 or total <= 0 or abs((quantity * rate) - total) > Decimal("0.02"): raise ValueError("Use valid BUY or SELL rows and ensure Total Amount equals Quantity × Rate.")
                        rows.append((transaction_type, stock_name, symbol, quantity, transaction_date, timestamp, rate, total))
                    if not rows: raise ValueError("The Excel file has no investment transactions.")
                    connection = db()
                    for transaction_type, stock_name, symbol, quantity, transaction_date, timestamp, rate, total in sorted(rows, key=lambda item: item[5]): connection.execute("INSERT INTO stock_transactions(company_id,transaction_type,stock_name,symbol,quantity,transaction_date,transaction_timestamp,rate,total_amount) VALUES(?,?,?,?,?,?,?,?,?)", (company["id"], transaction_type, stock_name, symbol, float(quantity), transaction_date, timestamp, float(rate), float(total)))
                    rebuild_stock_portfolio(connection, company["id"]); audit(company["id"], "Stock investment import", f"{len(rows)} transaction(s) imported using FIFO"); connection.commit(); flash(f"Imported {len(rows)} stock transaction(s) and updated FIFO holdings.", "success")
                except (ValueError, InvalidOperation) as error:
                    db().rollback(); flash("Investment import failed. " + str(error), "error")
                except sqlite3.Error:
                    db().rollback(); flash("Investment import could not be saved. Please try again.", "error")
                return redirect(url_for("investment_analysis"))
            frame = pd.read_excel(upload)
            required = ["Transaction_Type", "Stock_Name", "Symbol", "Quantity", "Purchase_Date", "Rate", "Total_Amount"]
            broker_columns = ["Name", "Date & time", "Side", "Qty", "Traded price", "Total value", "Segment"]
            if all(column in frame.columns for column in broker_columns):
                frame = frame.rename(columns={"Name": "Stock_Name", "Date & time": "Purchase_Date", "Side": "Transaction_Type", "Qty": "Quantity", "Traded price": "Rate", "Total value": "Total_Amount"})
                frame["Symbol"] = frame["Stock_Name"].astype(str).str.upper().str.replace(r"[^A-Z0-9]", "", regex=True).str[:24]
            elif any(column not in frame.columns for column in required): raise ValueError("Use the downloadable broker template or the standard investment template.")
            frame = frame[required].dropna(how="all").copy()
            if frame.empty or frame.isna().any().any(): raise ValueError("Every row must contain all required columns.")
            frame["Transaction_Type"] = frame["Transaction_Type"].astype(str).str.strip().str.upper()
            frame["Stock_Name"] = frame["Stock_Name"].astype(str).str.strip()
            frame["Symbol"] = frame["Symbol"].astype(str).str.strip().str.upper()
            frame["Purchase_Date"] = pd.to_datetime(frame["Purchase_Date"], errors="coerce", dayfirst=True)
            for column in ("Quantity", "Rate", "Total_Amount"): frame[column] = pd.to_numeric(frame[column].astype(str).str.replace(",", "", regex=False), errors="coerce")
            if frame.empty or frame.isna().any().any() or not frame["Transaction_Type"].isin(["BUY", "SELL"]).all() or (frame[["Quantity", "Rate", "Total_Amount"]] <= 0).any().any() or (frame["Stock_Name"].isin(["", "NAN", "NONE"])).any() or (frame["Symbol"].isin(["", "NAN", "NONE"])).any(): raise ValueError("Use BUY or SELL only and enter valid names, symbols, dates and positive amounts.")
            connection = db()
            for row in frame.sort_values("Purchase_Date", kind="stable").itertuples(index=False):
                quantity, rate, total = Decimal(str(row.Quantity)), Decimal(str(row.Rate)), Decimal(str(row.Total_Amount))
                if abs((quantity * rate) - total) > Decimal("0.02"): raise ValueError(f"Total Amount must equal Quantity × Rate for {row.Symbol}.")
                connection.execute("INSERT INTO stock_transactions(company_id,transaction_type,stock_name,symbol,quantity,transaction_date,transaction_timestamp,rate,total_amount) VALUES(?,?,?,?,?,?,?,?,?)", (company["id"], row.Transaction_Type, row.Stock_Name, row.Symbol, float(quantity), row.Purchase_Date.date().isoformat(), row.Purchase_Date.isoformat(), float(rate), float(total)))
            rebuild_stock_portfolio(connection, company["id"])
            audit(company["id"], "Stock investment import", f"{len(frame)} transaction(s) imported using FIFO")
            connection.commit(); flash(f"Imported {len(frame)} stock transaction(s) and updated FIFO holdings.", "success")
        except RuntimeError: flash("Investment import needs pandas. Install the updated requirements and restart Zedjer.", "error")
        except (ValueError, InvalidOperation) as error:
            db().rollback(); flash("Investment import failed. " + (str(error) or "Check the Excel template and holdings."), "error")
        except sqlite3.Error:
            db().rollback(); flash("Investment import could not be saved. Please try again.", "error")
        return redirect(url_for("investment_analysis"))
    today = date.today()
    holdings = db().execute("SELECT * FROM stock_holdings WHERE company_id=? ORDER BY symbol", (company["id"],)).fetchall()
    holding_rows = []
    for holding in holdings:
        lots = db().execute("SELECT * FROM stock_lots WHERE company_id=? AND symbol=? AND remaining_quantity>0 ORDER BY purchase_date,id", (company["id"], holding["symbol"])).fetchall()
        short_value = sum((Decimal(str(lot["remaining_quantity"])) * Decimal(str(lot["rate"])) for lot in lots if (today - date.fromisoformat(lot["purchase_date"])).days <= 365), Decimal("0"))
        investment_value = Decimal(str(holding["remaining_quantity"])) * Decimal(str(holding["average_cost"]))
        long_value = investment_value - short_value
        holding_rows.append({**dict(holding), "investment_value": float(investment_value), "short_value": float(short_value), "long_value": float(long_value)})
    portfolio_totals = {"value": sum(Decimal(str(row["investment_value"])) for row in holding_rows), "quantity": sum(Decimal(str(row["remaining_quantity"])) for row in holding_rows), "short_value": sum(Decimal(str(row["short_value"])) for row in holding_rows), "long_value": sum(Decimal(str(row["long_value"])) for row in holding_rows)}
    period_from = request.args.get("period_from") or company["financial_year_start"]
    period_to = request.args.get("period_to") or company["financial_year_end"]
    try:
        period_from, period_to = normalise_date(period_from), normalise_date(period_to)
        if period_from > period_to: raise ValueError
    except ValueError:
        period_from, period_to = company["financial_year_start"], company["financial_year_end"]
    detail_sql = """SELECT m.*,l.stock_name,l.purchase_date,m.quantity*m.fifo_cost_rate buy_value,m.quantity*m.sale_rate sale_value
      FROM stock_realized_matches m JOIN stock_lots l ON l.id=m.buy_lot_id
      WHERE m.company_id=? AND m.sale_date BETWEEN ? AND ? ORDER BY m.sale_date DESC,m.id DESC"""
    realized = db().execute(detail_sql, (company["id"], period_from, period_to)).fetchall()
    summary_sort = request.args.get("summary_sort", "name")
    summary_order = {"name": "l.stock_name ASC,m.tax_classification", "pnl_high": "realized_gain_loss DESC", "pnl_low": "realized_gain_loss ASC", "buy_high": "buy_value DESC", "sale_high": "sale_value DESC"}.get(summary_sort, "l.stock_name ASC,m.tax_classification")
    summary = db().execute(f"""SELECT m.symbol,l.stock_name,m.tax_classification,SUM(m.quantity) quantity,
      SUM(m.quantity*m.fifo_cost_rate) buy_value,SUM(m.quantity*m.sale_rate) sale_value,SUM(m.realized_gain_loss) realized_gain_loss
      FROM stock_realized_matches m JOIN stock_lots l ON l.id=m.buy_lot_id
      WHERE m.company_id=? AND m.sale_date BETWEEN ? AND ? GROUP BY m.symbol,l.stock_name,m.tax_classification
      ORDER BY {summary_order}""", (company["id"], period_from, period_to)).fetchall()
    summary = [{**dict(row), "pnl_percentage": float((Decimal(str(row["realized_gain_loss"])) / Decimal(str(row["buy_value"]))) * 100) if Decimal(str(row["buy_value"])) else 0} for row in summary]
    pnl_totals = {"short_term": sum(Decimal(str(row["realized_gain_loss"])) for row in summary if row["tax_classification"] == "Short-Term"), "long_term": sum(Decimal(str(row["realized_gain_loss"])) for row in summary if row["tax_classification"] == "Long-Term")}
    pnl_totals["total"] = pnl_totals["short_term"] + pnl_totals["long_term"]
    if request.args.get("export"):
        report_kind = request.args.get("report_kind", "summary")
        if report_kind == "detailed":
            export_rows = [[row["stock_name"], row["symbol"], display_date(row["purchase_date"]), display_date(row["sale_date"]), money(row["quantity"]), money(row["fifo_cost_rate"]), money(row["buy_value"]), money(row["sale_rate"]), money(row["sale_value"]), money(row["realized_gain_loss"]), row["holding_days"], row["tax_classification"]] for row in realized]
            result = export_report(company, "Stock P&L Detailed FIFO", ["Stock", "Symbol", "Buy date", "Sale date", "Quantity", "Buy price", "Buy value", "Sale price", "Sale value", "Realized P&L", "Holding days", "Classification"], export_rows, company["base_currency"])
        else:
            export_rows = [[row["stock_name"], row["tax_classification"], money(row["quantity"]), money(row["buy_value"]), money(row["sale_value"]), money(row["realized_gain_loss"]), money(row["pnl_percentage"]) + "%"] for row in summary]
            export_rows.append(["TOTAL", "", "", "", "", money(pnl_totals["total"]), ""])
            result = export_report(company, "Stock P&L Summary", ["Stock", "Classification", "Quantity sold", "Buy value", "Sale value", "Realized P&L", "P&L %"], export_rows, company["base_currency"])
        if result: return result
    return render_template("investment_analysis.html", holdings=holding_rows, portfolio_totals=portfolio_totals, realized=realized, summary=summary, summary_sort=summary_sort, pnl_totals=pnl_totals, period_from=period_from, period_to=period_to)


@app.route("/investment-analysis/sample")
def investment_analysis_sample():
    company = company_required()
    if not company: return redirect(url_for("companies_dashboard"))
    headers = ["Name", "Date & time", "Side", "Qty", "Traded price", "Total value", "Segment"]
    rows = [["Example Industries", "01 Jan 2026, 09:30:00 AM", "BUY", 10, 100, 1000, "Equity"], ["Example Industries", "15 Jul 2026, 11:15:00 AM", "SELL", 2, 120, 240, "Equity"]]
    output = BytesIO()
    try:
        import pandas as pd
        pd.DataFrame(rows, columns=headers).to_excel(output, index=False)
    except ModuleNotFoundError:
        from openpyxl import Workbook
        book = Workbook(); sheet = book.active; sheet.title = "Stock Investments"; sheet.append(headers)
        for row in rows: sheet.append(row)
        for column in sheet.columns: sheet.column_dimensions[column[0].column_letter].width = 19
        book.save(output)
    return Response(output.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=zedjer-stock-investment-template.xlsx"})


@app.route("/loan-amortization", methods=["GET", "POST"])
def loan_amortization():
    company = company_required()
    if not company: return redirect(url_for("companies_dashboard"))
    loans = db().execute("SELECT * FROM accounts WHERE company_id=? AND category='Liability' AND is_loan=1 ORDER BY code", (company["id"],)).fetchall()
    account_id = request.values.get("account_id", type=int) or (loans[0]["id"] if loans else None)
    if request.method == "POST":
        try:
            if account_id not in {loan["id"] for loan in loans}: raise ValueError
            if request.form.get("form_kind") == "profile":
                principal, annual_rate = Decimal(request.form.get("principal") or 0), Decimal(request.form.get("annual_rate") or 0)
                term = int(request.form.get("term_months") or 0)
                if principal < 0 or annual_rate < 0 or term < 0: raise ValueError
                db().execute("INSERT INTO loan_profiles(account_id,company_id,lender,principal,annual_rate,start_date,term_months,notes) VALUES(?,?,?,?,?,?,?,?) ON CONFLICT(account_id) DO UPDATE SET lender=excluded.lender,principal=excluded.principal,annual_rate=excluded.annual_rate,start_date=excluded.start_date,term_months=excluded.term_months,notes=excluded.notes", (account_id, company["id"], request.form.get("lender", "").strip(), float(principal), float(annual_rate), normalise_date(request.form.get("start_date") or company["financial_year_start"]), term, request.form.get("notes", "").strip()))
                audit(company["id"], "Loan details updated", str(account_id))
            else:
                principal, interest = Decimal(request.form.get("principal") or 0), Decimal(request.form.get("interest") or 0)
                if principal < 0 or interest < 0: raise ValueError
                db().execute("INSERT INTO loan_repayments(account_id,due_date,principal,interest,paid) VALUES(?,?,?,?,?)", (account_id, normalise_date(request.form["due_date"]), float(principal), float(interest), int(bool(request.form.get("paid")))))
                audit(company["id"], "Loan repayment schedule added", str(account_id))
            db().commit(); flash("Loan amortization details saved.", "success")
        except (ValueError, InvalidOperation): flash("Enter valid loan and repayment details.", "error")
        return redirect(url_for("loan_amortization", account_id=account_id))
    profile = db().execute("SELECT * FROM loan_profiles WHERE account_id=?", (account_id,)).fetchone() if account_id else None
    schedule = db().execute("SELECT * FROM loan_repayments WHERE account_id=? ORDER BY due_date,id", (account_id,)).fetchall() if account_id else []
    if account_id and allocate_loan_opening_schedule(company["id"], account_id):
        db().commit()
        schedule = db().execute("SELECT * FROM loan_repayments WHERE account_id=? ORDER BY due_date,id", (account_id,)).fetchall()
    balance = db().execute("SELECT COALESCE(SUM((l.credit-l.debit)*l.fx_rate),0) FROM journal_lines l JOIN journal_entries e ON e.id=l.entry_id WHERE e.company_id=? AND l.account_id=?", (company["id"], account_id)).fetchone()[0] if account_id else 0
    total_interest = sum(float(row["interest"] or 0) for row in schedule)
    if schedule:
        first = schedule[0]; opening_amount = float(first["opening_principal"] if first["opening_principal"] is not None else (profile["principal"] if profile else 0))
        monthly_payment = float(first["installment_amount"] if first["installment_amount"] is not None else float(first["principal"] or 0) + float(first["interest"] or 0))
        annual_rate = (float(first["interest"] or 0) * 12 / opening_amount * 100) if opening_amount else float(profile["annual_rate"] or 0) if profile else 0
        total_cost = sum(float(row["installment_amount"] if row["installment_amount"] is not None else float(row["principal"] or 0) + float(row["interest"] or 0)) for row in schedule)
        loan_summary = {"amount": opening_amount, "annual_rate": annual_rate, "years": len(schedule) / 12, "start_date": first["due_date"], "monthly_payment": monthly_payment, "payments": len(schedule), "total_interest": total_interest, "total_cost": total_cost}
    else:
        loan_summary = {"amount": float(profile["principal"] or 0) if profile else 0, "annual_rate": float(profile["annual_rate"] or 0) if profile else 0, "years": (int(profile["term_months"] or 0) / 12) if profile else 0, "start_date": profile["start_date"] if profile else None, "monthly_payment": 0, "payments": int(profile["term_months"] or 0) if profile else 0, "total_interest": 0, "total_cost": float(profile["principal"] or 0) if profile else 0}
    next_outstanding = next(((index, row) for index, row in enumerate(schedule, 1) if not row["paid"]), None)
    next_outstanding_id = next_outstanding[1]["id"] if next_outstanding else None
    next_installment_no = (next_outstanding[1]["installment_no"] or next_outstanding[0]) if next_outstanding else None
    return render_template("loan_amortization.html", loans=loans, account_id=account_id, profile=profile, schedule=schedule, balance=balance, loan_summary=loan_summary, next_outstanding_id=next_outstanding_id, next_installment_no=next_installment_no)


def allocate_loan_opening_schedule(company_id, account_id):
    """Mark installments already covered by the loan opening balance as paid."""
    opening = db().execute("SELECT effective_date,debit,credit FROM opening_balances WHERE company_id=? AND account_id=? ORDER BY effective_date DESC LIMIT 1", (company_id, account_id)).fetchone()
    schedule = db().execute("SELECT id,due_date,opening_principal,principal,paid FROM loan_repayments WHERE account_id=? ORDER BY due_date,id", (account_id,)).fetchall()
    if not opening or not schedule:
        return 0
    outstanding = max(0, float(opening["credit"] or 0) - float(opening["debit"] or 0))
    first_opening = next((float(row["opening_principal"]) for row in schedule if row["opening_principal"] is not None), 0)
    repaid_principal = max(0, first_opening - outstanding)
    paid_ids, allocated = set(), 0.0
    for row in schedule:
        # A schedule line before the opening date is historical. Also mark whole
        # installments already represented by the reduced opening principal.
        if row["due_date"] < opening["effective_date"]:
            if not row["paid"]: paid_ids.add(row["id"])
            allocated += float(row["principal"] or 0)
        elif allocated + float(row["principal"] or 0) <= repaid_principal + .01:
            if not row["paid"]: paid_ids.add(row["id"])
            allocated += float(row["principal"] or 0)
    if paid_ids:
        placeholders = ",".join("?" for _ in paid_ids)
        db().execute(f"UPDATE loan_repayments SET paid=1 WHERE account_id=? AND id IN ({placeholders})", [account_id, *paid_ids])
    return len(paid_ids)


@app.route("/loan-amortization/import", methods=["POST"])
def import_loan_schedule():
    company = company_required()
    if not company: return redirect(url_for("companies_dashboard"))
    account_id, upload = request.form.get("account_id", type=int), request.files.get("file")
    try:
        valid_loan = db().execute("SELECT 1 FROM accounts WHERE id=? AND company_id=? AND category='Liability' AND is_loan=1", (account_id, company["id"])).fetchone()
        if not valid_loan or not upload or not upload.filename.lower().endswith(".xlsx"): raise ValueError
        try: from openpyxl import load_workbook
        except ModuleNotFoundError: raise RuntimeError
        sheet = load_workbook(upload, data_only=True).active
        headers = [str(cell.value or "").strip().lower() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        aliases = {"inst no": "installment_no", "installment no": "installment_no", "due date": "due_date", "opening princ": "opening_principal", "opening principal": "opening_principal", "inst amt": "installment_amount", "installment amount": "installment_amount", "principal": "principal", "interest": "interest", "closing princ": "closing_principal", "closing principal": "closing_principal"}
        index = {aliases.get(header, header): position for position, header in enumerate(headers)}
        required = ("installment_no", "due_date", "opening_principal", "installment_amount", "principal", "interest", "closing_principal")
        if any(column not in index for column in required): raise ValueError
        lines = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(value not in (None, "") for value in row): continue
            installment_no = int(Decimal(str(row[index["installment_no"]])))
            due_date = normalise_date(row[index["due_date"]])
            opening, amount = Decimal(str(row[index["opening_principal"]] or 0)), Decimal(str(row[index["installment_amount"]] or 0))
            principal, interest = Decimal(str(row[index["principal"]] or 0)), Decimal(str(row[index["interest"]] or 0))
            closing = Decimal(str(row[index["closing_principal"]] or 0))
            if installment_no < 1 or min(opening, amount, principal, interest, closing) < 0 or abs(amount - principal - interest) > Decimal(".02"): raise ValueError
            lines.append((account_id, installment_no, due_date, float(opening), float(amount), float(principal), float(interest), float(closing)))
        if not lines: raise ValueError
        connection = db(); connection.execute("DELETE FROM loan_repayments WHERE account_id=?", (account_id,))
        connection.executemany("INSERT INTO loan_repayments(account_id,installment_no,due_date,opening_principal,installment_amount,principal,interest,closing_principal) VALUES(?,?,?,?,?,?,?,?)", lines)
        paid_count = allocate_loan_opening_schedule(company["id"], account_id)
        audit(company["id"], "Loan repayment schedule imported", f"Loan {account_id}: {len(lines)} installment(s), {paid_count} allocated to opening balance")
        connection.commit(); flash(f"Imported {len(lines)} repayment schedule line(s). {paid_count} historical installment(s) marked paid from the opening balance.", "success")
    except RuntimeError: flash("Loan schedule import needs openpyxl. Run: python -m pip install -r requirements.txt", "error")
    except (ValueError, InvalidOperation, IndexError):
        db().rollback(); flash("Import failed. Use the sample Excel columns and ensure installment amount equals principal plus interest.", "error")
    return redirect(url_for("loan_amortization", account_id=account_id))


@app.route("/loan-amortization/import/sample")
def loan_schedule_import_sample():
    company = company_required()
    if not company: return redirect(url_for("companies_dashboard"))
    try: from openpyxl import Workbook
    except ModuleNotFoundError: return Response("<h2>Excel sample setup needed</h2><p>Please run: <code>python -m pip install -r requirements.txt</code>.</p>", status=503, mimetype="text/html")
    book = Workbook(); sheet = book.active; sheet.title = "Repayment Schedule"
    sheet.append(["Inst No", "Due Date", "Opening Princ", "Inst Amt", "Principal", "Interest", "Closing Princ"])
    sheet.append([1, "04-06-2026", 23327, 3131, 2755, 376, 20572])
    sheet.append([2, "04-07-2026", 20572, 3131, 2800, 331, 17772])
    for column in sheet.columns: sheet.column_dimensions[column[0].column_letter].width = 18
    output = BytesIO(); book.save(output)
    return Response(output.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=zedjer-loan-repayment-schedule-sample.xlsx"})


@app.route("/master/users", methods=["GET", "POST"])
def user_authorisations():
    company = company_required()
    if not company: return redirect(url_for("companies_dashboard"))
    if request.method == "POST":
        try:
            username, role = request.form.get("username", "").strip(), request.form.get("role", "Accountant")
            password = request.form.get("password", "")
            if not username or role not in ("Owner", "Administrator", "Accountant", "Viewer"): raise ValueError
            existing = db().execute("SELECT id,password_hash FROM company_users WHERE company_id=? AND username=? COLLATE NOCASE", (company["id"], username)).fetchone()
            if not password and not existing: raise ValueError
            password_hash = generate_password_hash(password.casefold()) if password else existing["password_hash"]
            if existing:
                db().execute("UPDATE company_users SET username=?,contact_no=?,email=?,password_hash=?,role=?,active=? WHERE id=?", (username, request.form.get("contact_no", "").strip(), request.form.get("email", "").strip(), password_hash, role, int(bool(request.form.get("active"))), existing["id"]))
            else:
                db().execute("INSERT INTO company_users(company_id,username,contact_no,email,password_hash,role,active) VALUES(?,?,?,?,?,?,?)", (company["id"], username, request.form.get("contact_no", "").strip(), request.form.get("email", "").strip(), password_hash, role, int(bool(request.form.get("active")))))
            audit(company["id"], "User authorisation updated", f"{username}: {role}"); db().commit(); flash("User authorisation saved.", "success")
        except (ValueError, sqlite3.IntegrityError): flash("Enter a user name and valid access level.", "error")
        return redirect(url_for("user_authorisations"))
    users = db().execute("SELECT * FROM company_users WHERE company_id=? ORDER BY username", (company["id"],)).fetchall()
    return render_template("user_authorisations.html", users=users, is_owner=session.get("user_role") == "Owner")


@app.route("/master/users/<int:user_id>/edit", methods=["GET", "POST"])
def edit_user_authorisation(user_id):
    company = company_required()
    if not company: return redirect(url_for("companies_dashboard"))
    user = db().execute("SELECT * FROM company_users WHERE id=? AND company_id=?", (user_id, company["id"])).fetchone()
    if not user: return redirect(url_for("user_authorisations"))
    if request.method == "POST":
        try:
            username = request.form.get("username", "").strip()
            role = request.form.get("role", "Accountant")
            password = request.form.get("password", "")
            if not username or role not in ("Owner", "Administrator", "Accountant", "Viewer"): raise ValueError
            if db().execute("SELECT 1 FROM company_users WHERE company_id=? AND username=? COLLATE NOCASE AND id<>?", (company["id"], username, user_id)).fetchone(): raise ValueError
            password_hash = generate_password_hash(password.casefold()) if password else user["password_hash"]
            active = int(bool(request.form.get("active")))
            if company["auth_enabled"] and not active and not db().execute("SELECT 1 FROM company_users WHERE company_id=? AND id<>? AND active=1 AND password_hash IS NOT NULL AND trim(password_hash)<>''", (company["id"], user_id)).fetchone():
                raise ValueError
            db().execute("UPDATE company_users SET username=?,contact_no=?,email=?,password_hash=?,role=?,active=? WHERE id=? AND company_id=?", (username, request.form.get("contact_no", "").strip(), request.form.get("email", "").strip(), password_hash, role, active, user_id, company["id"]))
            audit(company["id"], "User authorisation edited", f"{user['username']} changed to {username}: {role}")
            if session.get("username") == user["username"]: session["username"] = username
            db().commit(); flash("User details updated.", "success")
            return redirect(url_for("user_authorisations"))
        except (ValueError, sqlite3.IntegrityError): flash("Enter valid user details. Keep at least one active user while sign-in is enabled.", "error")
    return render_template("edit_user_authorisation.html", user=user)


@app.route("/master/user-authorisations/setting", methods=["POST"])
def user_authorisation_setting():
    company = company_required()
    if not company: return redirect(url_for("companies_dashboard"))
    enabled = int(bool(request.form.get("auth_enabled")))
    if enabled and not db().execute("SELECT 1 FROM company_users WHERE company_id=? AND active=1 AND password_hash IS NOT NULL AND trim(password_hash)<>''", (company["id"],)).fetchone():
        flash("Add an active user with a login password before enabling user authorisation.", "error")
        return redirect(url_for("user_authorisations"))
    db().execute("UPDATE companies SET auth_enabled=? WHERE id=?", (enabled, company["id"]))
    audit(company["id"], "User authorisation setting", "Enabled" if enabled else "Disabled")
    db().commit()
    flash("User authorisation is now enabled." if enabled else "User authorisation is now disabled.", "success")
    return redirect(url_for("user_authorisations"))


@app.route("/companies/<int:company_id>/login", methods=["GET", "POST"])
def user_login(company_id):
    company = db().execute("SELECT * FROM companies WHERE id=? AND name <> 'Imported company'", (company_id,)).fetchone()
    if not company: return redirect(url_for("companies_dashboard"))
    if request.method == "POST":
        user = db().execute("SELECT * FROM company_users WHERE company_id=? AND username=? COLLATE NOCASE AND active=1", (company_id, request.form.get("username", "").strip())).fetchone()
        entered_password = request.form.get("password", "")
        if user and user["password_hash"] and (check_password_hash(user["password_hash"], entered_password.casefold()) or check_password_hash(user["password_hash"], entered_password)):
            session.permanent = True
            session["company_id"], session["username"], session["user_role"] = company_id, user["username"], user["role"]; audit(company_id, "User login", f"Role: {user['role']}"); db().commit(); return redirect(url_for("analysis"))
        flash("Invalid user name or password.", "error")
    return render_template("user_login.html", company=company)


@app.route("/logout")
def user_logout():
    company = active_company()
    if company: audit(company["id"], "User logout"); db().commit()
    session.clear()
    return redirect(url_for("sign_in"))


@app.route("/master/activity")
def activity_records():
    company = company_required()
    if not company: return redirect(url_for("companies_dashboard"))
    rows = db().execute("SELECT * FROM activity_log WHERE company_id=? ORDER BY id DESC LIMIT 500", (company["id"],)).fetchall()
    return render_template("activity_records.html", rows=rows)


@app.route("/current-user.json")
def current_user_json():
    """Small UI helper for the signed-in name shown in the application header."""
    user = signed_in_user()
    if not user:
        return {"name": "", "role": ""}, 401
    return {"name": user["display_name"], "role": session.get("user_role", "Zedjer user"), "is_admin": bool(user["is_admin"])}


@app.route("/admin/users")
def admin_users():
    user = signed_in_user()
    if not user or not user["is_admin"]:
        flash("Administrator access is required.", "error")
        return redirect(url_for("companies_dashboard"))
    users = db().execute("""SELECT au.*, COALESCE(GROUP_CONCAT(c.name, ' · '),'No companies assigned') AS company_names
        FROM app_users au LEFT JOIN company_access ca ON ca.app_user_id=au.id
        LEFT JOIN companies c ON c.id=ca.company_id AND c.name <> 'Imported company'
        GROUP BY au.id ORDER BY au.is_admin DESC, au.display_name COLLATE NOCASE""").fetchall()
    reset_requests = db().execute("""SELECT pr.*,au.display_name,au.email FROM password_reset_requests pr
        JOIN app_users au ON au.id=pr.app_user_id WHERE pr.status='Pending' ORDER BY pr.requested_at""").fetchall()
    return render_template("admin_users.html", users=users, reset_requests=reset_requests)


@app.route("/admin/users/<int:user_id>/reset-password", methods=["POST"])
def admin_reset_password(user_id):
    admin = signed_in_user()
    if not admin or not admin["is_admin"]:
        return redirect(url_for("sign_in"))
    password = request.form.get("password", "")
    confirm_password = request.form.get("confirm_password", "")
    if len(password) < 8 or password != confirm_password:
        flash("Use a matching new password of at least 8 characters.", "error")
        return redirect(url_for("admin_users"))
    user = db().execute("SELECT * FROM app_users WHERE id=?", (user_id,)).fetchone()
    if not user:
        flash("User was not found.", "error")
        return redirect(url_for("admin_users"))
    db().execute("UPDATE app_users SET password_hash=? WHERE id=?", (generate_password_hash(password), user_id))
    db().execute("UPDATE password_reset_requests SET status='Completed' WHERE app_user_id=? AND status='Pending'", (user_id,))
    db().commit()
    flash("Password updated for " + user["display_name"] + ".", "success")
    return redirect(url_for("admin_users"))


@app.route("/companies/select", methods=["POST"])
def select_company():
    company_id = request.form.get("company_id", type=int)
    if company_id and session.get("company_id") == company_id and session.get("username"):
        return redirect(url_for("analysis"))
    company = db().execute("SELECT * FROM companies WHERE id=? AND name <> 'Imported company'", (company_id,)).fetchone()
    if company:
        app_user = signed_in_user()
        access = user_company_access(app_user, company_id)
        if not access:
            flash("This account does not have access to the selected company.", "error")
            return redirect(url_for("companies_dashboard"))
        session.permanent = True
        session["company_id"], session["username"], session["user_role"] = company_id, app_user["display_name"], access["role"]
        audit(company_id, "Company selected", f"Email sign-in: {app_user['email']}")
        db().commit()
        flash("Company selected.", "success")
    return redirect(url_for("analysis"))


@app.route("/analysis")
def analysis():
    company = company_required()
    if not company: return redirect(url_for("companies_dashboard"))
    start, end = report_period(company)
    currency_code, factor = report_factor(company)
    factor = float(factor)
    region_id = request.args.get("region_tag_id", type=int)
    selected_analysis_categories = {int(value) for value in request.args.getlist("analysis_category_ids") if value.isdigit()}
    tag_id = request.args.get("tag_id", type=int)
    tagged_account_ids = {row[0] for row in db().execute("SELECT account_id FROM account_tag_links WHERE tag_id=?", (tag_id,)).fetchall()} if tag_id else set()
    compare_start = request.args.get("compare_from") or f"{int(start[:4])-1}{start[4:]}"; compare_end = request.args.get("compare_to") or f"{int(end[:4])-1}{end[4:]}"
    data = balances(company["id"], end); totals = {kind: sum(item["balance"] for item in data if item["category"] == kind) for kind in ("Asset", "Liability", "Equity", "Income", "Expense")}
    if tag_id:
        data = [item for item in data if item["id"] in tagged_account_ids]
        totals = {kind: sum(item["balance"] for item in data if item["category"] == kind) for kind in ("Asset", "Liability", "Equity", "Income", "Expense")}
    def period_categories(period_start, period_end):
        rows = db().execute("SELECT a.id,a.category,COALESCE(SUM((l.credit-l.debit)*l.fx_rate),0) amount FROM journal_lines l JOIN journal_entries e ON e.id=l.entry_id JOIN accounts a ON a.id=l.account_id WHERE e.company_id=? AND e.entry_date BETWEEN ? AND ? AND (? IS NULL OR e.region_tag_id=?) AND a.category IN ('Income','Expense') GROUP BY a.id,a.category", (company["id"], period_start, period_end, region_id, region_id)).fetchall()
        values = {kind: sum(row["amount"] for row in rows if row["category"] == kind and (not tag_id or row["id"] in tagged_account_ids)) for kind in ("Income", "Expense")}; return values["Income"], -values["Expense"]
    revenue, expenses = period_categories(start, end); previous_revenue, previous_expenses = period_categories(compare_start, compare_end)
    expense_rows = db().execute("SELECT a.id,a.name,COALESCE(SUM((l.debit-l.credit)*l.fx_rate),0) amount FROM journal_lines l JOIN journal_entries e ON e.id=l.entry_id JOIN accounts a ON a.id=l.account_id WHERE e.company_id=? AND e.entry_date BETWEEN ? AND ? AND (? IS NULL OR e.region_tag_id=?) AND a.category='Expense' GROUP BY a.id ORDER BY amount DESC", (company["id"], start, end, region_id, region_id)).fetchall()
    if tag_id: expense_rows = [row for row in expense_rows if row["id"] in tagged_account_ids]
    analysis_tag_rows = db().execute("SELECT c.id category_id,c.name category,t.name tag,COALESCE(SUM(CASE WHEN a.category='Income' THEN (l.credit-l.debit)*l.fx_rate ELSE (l.debit-l.credit)*l.fx_rate END),0) amount FROM journal_entries e JOIN journal_lines l ON l.entry_id=e.id JOIN accounts a ON a.id=l.account_id JOIN accounting_tags t ON t.id=e.accounting_tag_id JOIN analysis_categories c ON c.id=t.analysis_category_id WHERE e.company_id=? AND e.entry_date BETWEEN ? AND ? AND (? IS NULL OR e.region_tag_id=?) AND a.category IN ('Income','Expense') GROUP BY c.id,t.id ORDER BY c.name,t.name", (company["id"], start, end, region_id, region_id)).fetchall()
    if selected_analysis_categories: analysis_tag_rows = [row for row in analysis_tag_rows if row["category_id"] in selected_analysis_categories]
    cash_flow = db().execute("SELECT COALESCE(SUM(l.debit*l.fx_rate),0) inflow,COALESCE(SUM(l.credit*l.fx_rate),0) outflow FROM journal_lines l JOIN journal_entries e ON e.id=l.entry_id JOIN accounts a ON a.id=l.account_id WHERE e.company_id=? AND e.entry_date BETWEEN ? AND ? AND (? IS NULL OR e.region_tag_id=?) AND a.is_cash=1 AND a.cash_type IN ('Cash','Bank','Card')", (company["id"], start, end, region_id, region_id)).fetchone()
    def monthly_amounts(where_sql, value_sql, extra=()):
        query = f"SELECT substr(e.entry_date,1,7) month,COALESCE(SUM({value_sql}),0) amount FROM journal_lines l JOIN journal_entries e ON e.id=l.entry_id JOIN accounts a ON a.id=l.account_id WHERE e.company_id=? AND e.entry_date BETWEEN ? AND ? AND (? IS NULL OR e.region_tag_id=?) AND {where_sql} GROUP BY substr(e.entry_date,1,7) ORDER BY month"
        return {row["month"]: float(row["amount"]) * factor for row in db().execute(query, (company["id"], start, end, region_id, region_id, *extra)).fetchall()}
    def account_breakdown(where_sql, value_sql):
        query = f"SELECT a.name,COALESCE(SUM({value_sql}),0) amount FROM journal_lines l JOIN journal_entries e ON e.id=l.entry_id JOIN accounts a ON a.id=l.account_id WHERE e.company_id=? AND e.entry_date BETWEEN ? AND ? AND (? IS NULL OR e.region_tag_id=?) AND {where_sql} GROUP BY a.id,a.name HAVING ABS(SUM({value_sql})) > .005 ORDER BY ABS(amount) DESC"
        return [(row["name"], float(row["amount"]) * factor) for row in db().execute(query, (company["id"], start, end, region_id, region_id)).fetchall()]
    revenue_months = monthly_amounts("a.category='Income'", "(l.credit-l.debit)*l.fx_rate")
    expense_months = monthly_amounts("a.category='Expense'", "(l.debit-l.credit)*l.fx_rate")
    cash_months = monthly_amounts("a.is_cash=1 AND a.cash_type IN ('Cash','Bank','Card')", "(l.debit-l.credit)*l.fx_rate")
    profit_months = {month: revenue_months.get(month, 0) - expense_months.get(month, 0) for month in sorted(set(revenue_months) | set(expense_months))}
    revenue_breakdown = account_breakdown("a.category='Income'", "(l.credit-l.debit)*l.fx_rate")
    expense_breakdown = account_breakdown("a.category='Expense'", "(l.debit-l.credit)*l.fx_rate")
    cash_breakdown = account_breakdown("a.is_cash=1 AND a.cash_type IN ('Cash','Bank','Card')", "(l.debit-l.credit)*l.fx_rate")
    # Queries return base-currency amounts. Apply the selected dashboard currency
    # once here so every card and chart uses the same converted figures.
    revenue *= factor; expenses *= factor; previous_revenue *= factor; previous_expenses *= factor
    profit = revenue-expenses; previous_profit = previous_revenue-previous_expenses
    expense_values = [row["amount"] * factor for row in expense_rows]
    analysis_values = [row["amount"] * factor for row in analysis_tag_rows]
    cash_inflow, cash_outflow = cash_flow["inflow"] * factor, cash_flow["outflow"] * factor
    net_cashflow = cash_inflow-cash_outflow
    expense_rows = [{**dict(row), "amount": amount} for row, amount in zip(expense_rows, expense_values)]
    analysis_tag_rows = [{**dict(row), "amount": amount} for row, amount in zip(analysis_tag_rows, analysis_values)]
    cash_flow = {"inflow": cash_inflow, "outflow": cash_outflow}
    ratios = {"current": (totals["Asset"] / totals["Liability"]) if totals["Liability"] else 0, "margin": (profit / revenue * 100) if revenue else 0, "expense": (expenses / revenue * 100) if revenue else 0}
    tags = db().execute("SELECT * FROM accounting_tags WHERE company_id=? AND active=1 ORDER BY name", (company["id"],)).fetchall(); analysis_categories_list = db().execute("SELECT * FROM analysis_categories WHERE company_id=? ORDER BY name", (company["id"],)).fetchall()
    def drill(title, monthly, breakdown): return {"title": title, "months": list(monthly.keys()), "monthly": list(monthly.values()), "labels": [item[0] for item in breakdown], "breakdown": [item[1] for item in breakdown]}
    dashboard_drills = {"revenue": drill("Revenue analysis", revenue_months, revenue_breakdown), "expenses": drill("Expense analysis", expense_months, expense_breakdown), "profit": drill("Net profit analysis", profit_months, [("Revenue", revenue), ("Expenses", expenses)]), "cashflow": drill("Net cash flow analysis", cash_months, cash_breakdown)}
    return render_template("analysis.html", totals=totals, profit=profit, ratios=ratios, start=start, end=end, compare_start=compare_start, compare_end=compare_end, revenue=revenue, expenses=expenses, previous_revenue=previous_revenue, previous_expenses=previous_expenses, previous_profit=previous_profit, expense_labels=[r['name'] for r in expense_rows], expense_values=[r['amount'] for r in expense_rows], analysis_labels=[f"{r['category']} · {r['tag']}" for r in analysis_tag_rows], analysis_values=[r['amount'] for r in analysis_tag_rows], cash_inflow=cash_flow["inflow"], cash_outflow=cash_flow["outflow"], net_cashflow=cash_flow["inflow"]-cash_flow["outflow"], tags=tags, tag_id=tag_id, analysis_categories=analysis_categories_list, selected_analysis_categories=selected_analysis_categories, dashboard_drills=dashboard_drills)


@app.route("/currencies", methods=["GET", "POST"])
def currencies():
    company = company_required()
    if not company: return redirect(url_for("companies_dashboard"))
    if request.method == "POST":
        try:
            code = request.form.get("code", "").upper().strip(); name = request.form.get("name", "").strip(); rate = Decimal(request.form.get("rate_to_base") or 0)
            if len(code) != 3 or not code.isalpha() or not name or rate <= 0: raise ValueError
            if code == company["base_currency"]: rate = Decimal(1)
            db().execute("INSERT INTO currencies(company_id,code,name,rate_to_base) VALUES(?,?,?,?) ON CONFLICT(company_id,code) DO UPDATE SET name=excluded.name,rate_to_base=excluded.rate_to_base,active=1", (company["id"], code, name, float(rate))); db().commit(); flash("Currency and exchange rate saved.", "success")
        except (ValueError, InvalidOperation): flash("Enter a valid three-letter currency and positive rate.", "error")
        return redirect(url_for("currencies"))
    currencies_list = currency_list(company)
    edit_code = request.args.get("edit", "").upper().strip()
    editing = next((currency for currency in currencies_list if currency["code"] == edit_code), None)
    return render_template("currencies.html", currencies=currencies_list, editing=editing)


@app.route("/account-subgroups", methods=["GET", "POST"])
def account_subgroups():
    company = company_required()
    if not company: return redirect(url_for("companies_dashboard"))
    if request.method == "POST":
        try:
            category = request.form.get("category", ""); name = request.form.get("name", "").strip()
            if category not in ("Asset", "Liability", "Equity", "Income", "Expense") or not name: raise ValueError
            db().execute("INSERT INTO account_subgroups(company_id,category,name) VALUES(?,?,?)", (company["id"], category, name)); db().commit(); flash("Subgroup added.", "success")
        except (ValueError, sqlite3.IntegrityError): flash("Use a unique subgroup name within its main group.", "error")
        return redirect(url_for("account_subgroups"))
    subgroups = db().execute("SELECT * FROM account_subgroups WHERE company_id=? ORDER BY category,name", (company["id"],)).fetchall()
    return render_template("account_subgroups.html", subgroups=subgroups)


@app.route("/accounting-tags", methods=["GET", "POST"])
def accounting_tags():
    company = company_required()
    if not company: return redirect(url_for("companies_dashboard"))
    if request.method == "POST":
        try:
            name = request.form.get("name", "").strip(); category_id = request.form.get("analysis_category_id", type=int)
            if not name or not category_id or not db().execute("SELECT 1 FROM analysis_categories WHERE id=? AND company_id=?", (category_id, company["id"])).fetchone(): raise ValueError
            db().execute("INSERT INTO accounting_tags(company_id,name,analysis_category_id) VALUES(?,?,?)", (company["id"], name, category_id)); audit(company["id"], "Analysis tag created", name); db().commit(); flash("Analysis tag created.", "success")
        except (ValueError, sqlite3.IntegrityError): flash("Use a unique tag name.", "error")
        return redirect(url_for("accounting_tags"))
    tags = db().execute("SELECT t.*,c.name category_name,COUNT(l.account_id) account_count FROM accounting_tags t LEFT JOIN analysis_categories c ON c.id=t.analysis_category_id LEFT JOIN account_tag_links l ON l.tag_id=t.id WHERE t.company_id=? GROUP BY t.id ORDER BY c.name,t.name", (company["id"],)).fetchall()
    categories = db().execute("SELECT * FROM analysis_categories WHERE company_id=? ORDER BY name", (company["id"],)).fetchall()
    return render_template("accounting_tags.html", tags=tags, categories=categories)


@app.route("/accounting-tags.json")
def accounting_tags_json():
    company = company_required()
    if not company: return {"tags": []}, 403
    tags = db().execute("SELECT id,name,color FROM accounting_tags WHERE company_id=? AND active=1 ORDER BY name", (company["id"],)).fetchall()
    return {"tags": [dict(tag) for tag in tags]}


@app.route("/transactions/<int:entry_id>/dimensions.json")
def transaction_dimensions_for_entry(entry_id):
    company = company_required()
    entry = db().execute("SELECT accounting_tag_id,region_tag_id,party,payment_mode FROM journal_entries WHERE id=? AND company_id=?", (entry_id, company["id"])).fetchone() if company else None
    return {"accounting_tag_id": entry["accounting_tag_id"] if entry else None, "region_tag_id": entry["region_tag_id"] if entry else None, "party": entry["party"] if entry else "", "payment_mode": entry["payment_mode"] if entry else ""}


@app.route("/analysis-categories", methods=["GET", "POST"])
def analysis_categories():
    company = company_required()
    if not company: return redirect(url_for("companies_dashboard"))
    if request.method == "POST":
        try:
            name = request.form.get("name", "").strip()
            if not name: raise ValueError
            db().execute("INSERT INTO analysis_categories(company_id,name) VALUES(?,?)", (company["id"], name)); db().commit(); flash("Analysis category added.", "success")
        except (ValueError, sqlite3.IntegrityError): flash("Use a unique analysis category name.", "error")
        return redirect(url_for("analysis_categories"))
    categories = db().execute("SELECT * FROM analysis_categories WHERE company_id=? ORDER BY name", (company["id"],)).fetchall()
    return render_template("analysis_categories.html", categories=categories)


@app.route("/analysis-categories/<int:category_id>/edit", methods=["GET", "POST"])
def edit_analysis_category(category_id):
    company = company_required()
    category = db().execute("SELECT * FROM analysis_categories WHERE id=? AND company_id=?", (category_id, company["id"])).fetchone() if company else None
    if not category: return redirect(url_for("accounting_tags"))
    if request.method == "POST":
        try:
            name = request.form.get("name", "").strip()
            if not name: raise ValueError
            db().execute("UPDATE analysis_categories SET name=? WHERE id=?", (name, category_id)); db().commit(); flash("Analysis category updated.", "success"); return redirect(url_for("accounting_tags"))
        except (ValueError, sqlite3.IntegrityError): flash("Use a unique category name.", "error")
    return render_template("edit_analysis_category.html", category=category)


@app.route("/analysis-tags/<int:tag_id>/edit", methods=["GET", "POST"])
def edit_analysis_tag(tag_id):
    company = company_required()
    tag = db().execute("SELECT * FROM accounting_tags WHERE id=? AND company_id=?", (tag_id, company["id"])).fetchone() if company else None
    if not tag: return redirect(url_for("accounting_tags"))
    categories = db().execute("SELECT * FROM analysis_categories WHERE company_id=? ORDER BY name", (company["id"],)).fetchall()
    if request.method == "POST":
        try:
            name, category_id = request.form.get("name", "").strip(), request.form.get("analysis_category_id", type=int)
            if not name or not db().execute("SELECT 1 FROM analysis_categories WHERE id=? AND company_id=?", (category_id, company["id"])).fetchone(): raise ValueError
            db().execute("UPDATE accounting_tags SET name=?,analysis_category_id=? WHERE id=?", (name, category_id, tag_id)); db().commit(); flash("Analysis tag updated.", "success"); return redirect(url_for("accounting_tags"))
        except (ValueError, sqlite3.IntegrityError): flash("Use a unique tag name and category.", "error")
    return render_template("edit_analysis_tag.html", tag=tag, categories=categories)


@app.route("/analysis-categories/<int:category_id>/delete", methods=["POST"])
def delete_analysis_category(category_id):
    company = company_required()
    if not company: return redirect(url_for("companies_dashboard"))
    if db().execute("SELECT 1 FROM accounting_tags WHERE company_id=? AND analysis_category_id=?", (company["id"], category_id)).fetchone():
        flash("Delete or move its Analysis Tags before deleting this category.", "error")
    else:
        db().execute("DELETE FROM analysis_categories WHERE id=? AND company_id=?", (category_id, company["id"])); db().commit(); flash("Analysis category deleted.", "success")
    return redirect(url_for("accounting_tags"))


@app.route("/analysis-tags/<int:tag_id>/delete", methods=["POST"])
def delete_analysis_tag(tag_id):
    company = company_required()
    if not company: return redirect(url_for("companies_dashboard"))
    db().execute("DELETE FROM accounting_tags WHERE id=? AND company_id=?", (tag_id, company["id"])); db().commit(); flash("Analysis tag deleted.", "success")
    return redirect(url_for("accounting_tags"))


@app.route("/region-tags", methods=["GET", "POST"])
def region_tags():
    company = company_required()
    if not company: return redirect(url_for("companies_dashboard"))
    if request.method == "POST":
        enabled = int(bool(request.form.get("regions_enabled")))
        db().execute("UPDATE company_settings SET regions_enabled=? WHERE company_id=?", (enabled, company["id"]))
        name = request.form.get("name", "").strip()
        if name:
            try: db().execute("INSERT INTO region_tags(company_id,name) VALUES(?,?)", (company["id"], name))
            except sqlite3.IntegrityError: flash("That region already exists.", "error")
        db().commit(); flash("Region settings saved.", "success")
        return redirect(url_for("region_tags"))
    setting = db().execute("SELECT * FROM company_settings WHERE company_id=?", (company["id"],)).fetchone()
    regions = db().execute("SELECT * FROM region_tags WHERE company_id=? AND active=1 ORDER BY name", (company["id"],)).fetchall()
    return render_template("region_tags.html", setting=setting, regions=regions)


@app.route("/transaction-dimensions.json")
def transaction_dimensions_json():
    company = company_required()
    if not company: return {"tags": [], "regions_enabled": False, "regions": []}, 403
    tags = db().execute("SELECT id,name FROM accounting_tags WHERE company_id=? AND active=1 ORDER BY name", (company["id"],)).fetchall()
    categories = db().execute("SELECT id,name FROM analysis_categories WHERE company_id=? ORDER BY name", (company["id"],)).fetchall()
    setting = db().execute("SELECT regions_enabled FROM company_settings WHERE company_id=?", (company["id"],)).fetchone()
    regions = db().execute("SELECT id,name FROM region_tags WHERE company_id=? AND active=1 ORDER BY name", (company["id"],)).fetchall()
    return {"tags": [dict(item) for item in tags], "categories": [dict(item) for item in categories], "regions_enabled": bool(setting and setting["regions_enabled"]), "regions": [dict(item) for item in regions]}


@app.route("/accounts/<int:account_id>/tags.json")
def account_tags_json(account_id):
    company = company_required()
    if not company: return {"tag_ids": []}, 403
    rows = db().execute("SELECT DISTINCT t.id FROM accounting_tags t JOIN accounts a ON a.company_id=t.company_id LEFT JOIN account_tag_links l ON l.tag_id=t.id AND l.account_id=a.id LEFT JOIN account_analysis_categories c ON c.account_id=a.id WHERE a.id=? AND a.company_id=? AND (l.tag_id IS NOT NULL OR c.category_id=t.analysis_category_id)", (account_id, company["id"])).fetchall()
    return {"tag_ids": [row[0] for row in rows]}


@app.route("/accounts/<int:account_id>/loan.json")
def account_loan_json(account_id):
    company = company_required()
    account = db().execute("SELECT is_loan FROM accounts WHERE id=? AND company_id=?", (account_id, company["id"])).fetchone() if company else None
    return {"is_loan": bool(account and account["is_loan"])}


@app.route("/accounts/<int:account_id>/analysis-categories.json")
def account_analysis_categories_json(account_id):
    company = company_required()
    if not company: return {"category_ids": []}, 403
    rows = db().execute("SELECT c.category_id FROM account_analysis_categories c JOIN accounts a ON a.id=c.account_id WHERE a.id=? AND a.company_id=?", (account_id, company["id"])).fetchall()
    return {"category_ids": [row[0] for row in rows]}


@app.route("/companies/<int:company_id>/delete", methods=["GET", "POST"])
def delete_company(company_id):
    company = db().execute("SELECT * FROM companies WHERE id=? AND name <> 'Imported company'", (company_id,)).fetchone()
    if not company:
        return redirect(url_for("companies_dashboard"))
    if request.method == "POST":
        if request.form.get("confirm") == "yes":
            connection = db()
            connection.execute("DELETE FROM journal_lines WHERE entry_id IN (SELECT id FROM journal_entries WHERE company_id=?)", (company_id,))
            connection.execute("DELETE FROM companies WHERE id=?", (company_id,)); connection.commit()
            if session.get("company_id") == company_id: session.pop("company_id", None)
            flash("Company and all of its records were deleted.", "success")
        return redirect(url_for("companies_dashboard"))
    return render_template("delete_company.html", company=company)


@app.route("/transactions", methods=["GET", "POST"])
def transactions():
    company = company_required()
    if not company: return redirect(url_for("companies_dashboard"))
    accounts = db().execute("SELECT * FROM accounts WHERE company_id=? AND active=1 ORDER BY code", (company["id"],)).fetchall()
    if request.method == "POST":
        try:
            entry_date = normalise_date(request.form["entry_date"])
            if not company["financial_year_start"] <= entry_date <= company["financial_year_end"]: raise ValueError
            document_type = request.form.get("document_type", "Journal")
            if document_type not in DOCUMENT_TYPES: raise ValueError
            series = document_series(company["id"], document_type)
            document_no = suggested_document_no(company, document_type) if series["number_mode"] == "automatic" else request.form.get("document_no", "").strip()
            if not document_no or duplicate_document_number(company["id"], document_type, document_no): raise ValueError
            lines = []
            raw = zip(request.form.getlist("account_id[]"), request.form.getlist("description[]"), request.form.getlist("debit[]"), request.form.getlist("credit[]"), request.form.getlist("currency[]"), request.form.getlist("fx_rate[]"))
            for account, description, debit, credit, code, rate in raw:
                debit, credit, rate = Decimal(debit or 0), Decimal(credit or 0), Decimal(rate or 1)
                if debit < 0 or credit < 0 or rate <= 0 or (debit and credit) or not code.strip() or (debit or credit) and not db().execute("SELECT 1 FROM accounts WHERE id=? AND company_id=?", (int(account), company["id"])).fetchone(): raise ValueError
                if debit or credit: lines.append((int(account), description.strip(), float(debit), float(credit), code.upper().strip(), float(rate)))
            if len(lines) < 2 or abs(sum(Decimal(str(x[2]))*Decimal(str(x[5])) for x in lines)-sum(Decimal(str(x[3]))*Decimal(str(x[5])) for x in lines)) > Decimal(".005"): raise ValueError
        except (KeyError, ValueError, InvalidOperation): flash("Use valid accounts, balance both sides, and select a date in this company’s financial year.", "error")
        else:
            try:
                payment_mode = request.form.get("payment_mode", "").strip() if document_type in ("Payments", "Receipts") else ""
                if payment_mode not in ("", "Cash", "Bank", "Card", "Other"): raise ValueError
                tag_value = next((value for value in request.form.getlist("accounting_tag_id[]") if value), "")
                tag_id, region_id = (int(tag_value) if tag_value.isdigit() else None), request.form.get("region_tag_id", type=int)
                connection = db(); entry = connection.execute("INSERT INTO journal_entries(company_id,entry_date,document_type,document_no,reference,memo,payment_mode,party,accounting_tag_id,region_tag_id) VALUES(?,?,?,?,?,?,?,?,?,?)", (company["id"], entry_date, document_type, document_no, request.form.get("reference", "").strip(), request.form.get("memo", "").strip(), payment_mode, request.form.get("party", "").strip(), tag_id, region_id))
                connection.executemany("INSERT INTO journal_lines(entry_id,account_id,description,debit,credit,currency,fx_rate) VALUES(?,?,?,?,?,?,?)", [(entry.lastrowid, *line) for line in lines])
                if series["number_mode"] == "automatic": connection.execute("UPDATE document_series SET next_number=next_number+1 WHERE company_id=? AND document_type=?", (company["id"], document_type))
                audit(company["id"], "Transaction created", f"{document_type} {document_no}")
                connection.commit(); flash("Transaction posted.", "success"); return redirect(url_for("transaction_detail", entry_id=entry.lastrowid))
            except sqlite3.IntegrityError:
                db().rollback(); flash("That document number already exists for this company.", "error")
            except sqlite3.Error:
                db().rollback(); flash("Unable to save this transaction. Refresh the app and try again.", "error")
    entries = db().execute("SELECT * FROM journal_entries WHERE company_id=? ORDER BY entry_date DESC,id DESC LIMIT 20", (company["id"],)).fetchall()
    series_data = {kind: {"mode": document_series(company["id"], kind)["number_mode"], "suggested": suggested_document_no(company, kind)} for kind in DOCUMENT_TYPES}
    return render_template("transactions.html", accounts=accounts, entries=entries, document_types=DOCUMENT_TYPES, suggested_document_no=suggested_document_no(company), series_data=series_data)


@app.route("/transactions/<int:entry_id>")
def transaction_detail(entry_id):
    company = company_required()
    if not company: return redirect(url_for("companies_dashboard"))
    entry = db().execute("SELECT * FROM journal_entries WHERE id=? AND company_id=?", (entry_id, company["id"])).fetchone()
    if not entry:
        flash("Transaction not found.", "error")
        return redirect(url_for("transactions"))
    lines = db().execute("SELECT l.*,a.code,a.name,l.debit*l.fx_rate base_debit,l.credit*l.fx_rate base_credit FROM journal_lines l JOIN accounts a ON a.id=l.account_id WHERE l.entry_id=? ORDER BY l.id", (entry_id,)).fetchall()
    return render_template("transaction_detail.html", entry=entry, lines=lines)


@app.route("/transactions/<int:entry_id>/edit", methods=["GET", "POST"])
def edit_transaction(entry_id):
    company = company_required()
    if not company: return redirect(url_for("companies_dashboard"))
    entry = db().execute("SELECT * FROM journal_entries WHERE id=? AND company_id=?", (entry_id, company["id"])).fetchone()
    if not entry: return redirect(url_for("transactions"))
    accounts = db().execute("SELECT * FROM accounts WHERE company_id=? AND active=1 ORDER BY code", (company["id"],)).fetchall()
    lines = db().execute("SELECT * FROM journal_lines WHERE entry_id=? ORDER BY id", (entry_id,)).fetchall()
    if request.method == "POST":
        try:
            if company["edit_pin"] and request.form.get("edit_pin", "") != company["edit_pin"]: raise PermissionError
            entry_date = normalise_date(request.form["entry_date"])
            document_type = request.form.get("document_type", "Journal")
            document_no = request.form.get("document_no", "").strip()
            type_changed = document_type != entry["document_type"]
            series = document_series(company["id"], document_type) if document_type in DOCUMENT_TYPES else None
            if type_changed and series and series["number_mode"] == "automatic":
                document_no = suggested_document_no(company, document_type)
            number_changed = document_type != entry["document_type"] or document_no != entry["document_no"]
            if document_type not in DOCUMENT_TYPES or not document_no or (number_changed and duplicate_document_number(company["id"], document_type, document_no, entry_id)) or not company["financial_year_start"] <= entry_date <= company["financial_year_end"]: raise ValueError
            updated_lines = []
            for account, description, debit, credit, currency, rate in zip(request.form.getlist("account_id[]"), request.form.getlist("description[]"), request.form.getlist("debit[]"), request.form.getlist("credit[]"), request.form.getlist("currency[]"), request.form.getlist("fx_rate[]")):
                debit, credit, rate = Decimal(debit or 0), Decimal(credit or 0), Decimal(rate or 1)
                if int(account) not in [row["id"] for row in accounts] or debit < 0 or credit < 0 or rate <= 0 or (debit and credit): raise ValueError
                if debit or credit: updated_lines.append((int(account), description.strip(), float(debit), float(credit), currency.upper().strip(), float(rate)))
            if len(updated_lines) < 2 or abs(sum(Decimal(str(x[2]))*Decimal(str(x[5])) for x in updated_lines)-sum(Decimal(str(x[3]))*Decimal(str(x[5])) for x in updated_lines)) > Decimal(".005"): raise ValueError
            payment_mode = request.form.get("payment_mode", "").strip() if document_type in ("Payments", "Receipts") else ""
            if payment_mode not in ("", "Cash", "Bank", "Card", "Other"): raise ValueError
            tag_value = next((value for value in request.form.getlist("accounting_tag_id[]") if value), "")
            tag_id, region_id = (int(tag_value) if tag_value.isdigit() else None), request.form.get("region_tag_id", type=int)
            connection = db(); connection.execute("UPDATE journal_entries SET entry_date=?,document_type=?,document_no=?,reference=?,memo=?,payment_mode=?,party=?,accounting_tag_id=?,region_tag_id=? WHERE id=?", (entry_date, document_type, document_no, request.form.get("reference", "").strip(), request.form.get("memo", "").strip(), payment_mode, request.form.get("party", "").strip(), tag_id, region_id, entry_id)); connection.execute("UPDATE bank_statement_lines SET matched_journal_line_id=NULL WHERE matched_journal_line_id IN (SELECT id FROM journal_lines WHERE entry_id=?)", (entry_id,)); connection.execute("DELETE FROM journal_lines WHERE entry_id=?", (entry_id,)); connection.executemany("INSERT INTO journal_lines(entry_id,account_id,description,debit,credit,currency,fx_rate) VALUES(?,?,?,?,?,?,?)", [(entry_id, *line) for line in updated_lines]);
            if type_changed and series and series["number_mode"] == "automatic": connection.execute("UPDATE document_series SET next_number=next_number+1 WHERE company_id=? AND document_type=?", (company["id"], document_type))
            account_names = {row["id"]: row["name"] for row in accounts}
            def line_summary(items):
                return ", ".join(f"{account_names.get(int(item[0] if isinstance(item, tuple) else item['account_id']), 'Account')} Dr {float(item[2] if isinstance(item, tuple) else item['debit']):.2f} Cr {float(item[3] if isinstance(item, tuple) else item['credit']):.2f} {item[4] if isinstance(item, tuple) else item['currency']} @ {float(item[5] if isinstance(item, tuple) else item['fx_rate']):.4f}" for item in items)
            changes = []
            comparisons = (("Date", entry["entry_date"], entry_date), ("Document type", entry["document_type"], document_type), ("Document no.", entry["document_no"], document_no), ("Reference", entry["reference"] or "-", request.form.get("reference", "").strip() or "-"), ("Narration", entry["memo"] or "-", request.form.get("memo", "").strip() or "-"), ("Party", entry["party"] or "-", request.form.get("party", "").strip() or "-"), ("Payment/receipt mode", entry["payment_mode"] or "-", payment_mode or "-"), ("Region", entry["region_tag_id"] or "-", region_id or "-"))
            changes.extend(f"{label}: {before} → {after}" for label, before, after in comparisons if str(before) != str(after))
            if line_summary(lines) != line_summary(updated_lines): changes.append(f"Ledger lines: [{line_summary(lines)}] → [{line_summary(updated_lines)}]")
            audit(company["id"], "Transaction edited", f"{entry['document_no']} | " + ("; ".join(changes) or "No field values changed."))
            connection.commit()
            flash("Transaction updated.", "success"); return redirect(url_for("transaction_detail", entry_id=entry_id))
        except PermissionError: flash("Enter the company 4-digit edit PIN to edit this transaction.", "error")
        except (ValueError, InvalidOperation, sqlite3.IntegrityError): db().rollback(); flash("Unable to save: keep debit and credit equal, use valid accounts, and use a document number that is not already used for this document type.", "error")
    return render_template("edit_transaction.html", entry=entry, lines=lines, accounts=accounts, document_types=DOCUMENT_TYPES)


@app.route("/transactions/<int:entry_id>/delete", methods=["POST"])
def delete_transaction(entry_id):
    company = company_required()
    if not company: return redirect(url_for("companies_dashboard"))
    entry = db().execute("SELECT id,document_type,document_no FROM journal_entries WHERE id=? AND company_id=?", (entry_id, company["id"])).fetchone()
    if not entry:
        flash("Transaction not found.", "error")
    elif company["edit_pin"] and request.form.get("edit_pin", "") != company["edit_pin"]:
        flash("Enter the correct 4-digit company PIN to delete a transaction.", "error")
    else:
        db().execute("DELETE FROM journal_entries WHERE id=?", (entry_id,)); audit(company["id"], "Transaction deleted", f"{entry['document_type']} {entry['document_no']}"); db().commit(); flash("Transaction deleted.", "success")
    return_to = request.form.get("return_to", "")
    return redirect(return_to if return_to.startswith("/") else url_for("transactions"))


@app.route("/transactions/delete-selected", methods=["POST"])
def delete_selected_transactions():
    company = company_required()
    if not company: return redirect(url_for("companies_dashboard"))
    entry_ids = list({int(value) for value in request.form.getlist("entry_ids") if value.isdigit()})
    if not entry_ids:
        flash("Select at least one transaction.", "error")
    elif company["edit_pin"] and request.form.get("edit_pin", "") != company["edit_pin"]:
        flash("Enter the correct 4-digit company PIN to delete transactions.", "error")
    else:
        placeholders = ",".join("?" for _ in entry_ids)
        db().execute(f"DELETE FROM journal_entries WHERE company_id=? AND id IN ({placeholders})", [company["id"], *entry_ids]); audit(company["id"], "Transactions deleted", f"{len(entry_ids)} transaction(s)"); db().commit()
        flash(f"Deleted {len(entry_ids)} transaction(s).", "success")
    return_to = request.form.get("return_to", "")
    return redirect(return_to if return_to.startswith("/") else url_for("transactions"))


@app.route("/transactions/import", methods=["GET", "POST"])
def import_transactions():
    company = company_required()
    if not company: return redirect(url_for("companies_dashboard"))
    if request.method == "POST":
        upload = request.files.get("file")
        try:
            if not upload or not upload.filename.lower().endswith(".xlsx"): raise ValueError
            try: from openpyxl import load_workbook
            except ModuleNotFoundError: raise RuntimeError
            sheet = load_workbook(upload, data_only=True).active
            headers = [str(c.value or "").strip().lower() for c in next(sheet.iter_rows(min_row=1, max_row=1))]
            required = ["date", "debit", "credit", "currency", "fx rate"]
            if any(name not in headers for name in required): raise ValueError
            account_column = "account ledger" if "account ledger" in headers else "account name" if "account name" in headers else "account code" if "account code" in headers else None
            if not account_column: raise ValueError
            index = {name: headers.index(name) for name in headers}; account_rows = db().execute("SELECT id,code,name FROM accounts WHERE company_id=?", (company["id"],)).fetchall(); accounts = {r["name"]: r["id"] for r in account_rows} if account_column != "account code" else {r["code"]: r["id"] for r in account_rows}
            entries = {}
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not any(row): continue
                entry_date = normalise_date(row[index["date"]]); ledger = str(row[index[account_column]]).strip(); debit = Decimal(str(row[index["debit"]] or 0)); credit = Decimal(str(row[index["credit"]] or 0)); currency = str(row[index["currency"]] or company["base_currency"]).upper(); rate = Decimal(str(row[index["fx rate"]] or 1))
                if ledger not in accounts or debit < 0 or credit < 0 or (debit and credit) or rate <= 0: raise ValueError
                imported_no = str(row[index["document no"]] or "").strip() if "document no" in index else ""
                document_type = str(row[index["document type"]] or "Journal").title() if "document type" in index else "Journal"
                if document_type not in DOCUMENT_TYPES: raise ValueError
                key = (document_type, imported_no, entry_date, str(row[index.get("reference", -1)] or "") if "reference" in index else "", str(row[index.get("memo", -1)] or "") if "memo" in index else "")
                entries.setdefault(key, []).append((accounts[ledger], str(row[index.get("description", -1)] or "") if "description" in index else "", float(debit), float(credit), currency, float(rate)))
            connection = db()
            series_numbers = {kind: int(document_series(company["id"], kind)["next_number"]) for kind in DOCUMENT_TYPES}
            for (document_type, imported_no, entry_date, reference, memo), lines in entries.items():
                if not company["financial_year_start"] <= entry_date <= company["financial_year_end"]: raise ValueError
                if len(lines) < 2 or abs(sum(Decimal(str(x[2]))*Decimal(str(x[5])) for x in lines)-sum(Decimal(str(x[3]))*Decimal(str(x[5])) for x in lines)) > Decimal(".005"): raise ValueError
                series = document_series(company["id"], document_type)
                if series["number_mode"] == "manual" and not imported_no: raise ValueError
                document_no = imported_no or f"{series['prefix'] or ''}{series_numbers[document_type]:05d}"
                if not document_no: raise ValueError
                entry = connection.execute("INSERT INTO journal_entries(company_id,entry_date,document_type,document_no,reference,memo) VALUES(?,?,?,?,?,?)", (company["id"], entry_date, document_type, document_no, reference, memo)); connection.executemany("INSERT INTO journal_lines(entry_id,account_id,description,debit,credit,currency,fx_rate) VALUES(?,?,?,?,?,?,?)", [(entry.lastrowid, *line) for line in lines])
                if series["number_mode"] == "automatic" and not imported_no: series_numbers[document_type] += 1
            for document_type, next_number in series_numbers.items(): connection.execute("UPDATE document_series SET next_number=? WHERE company_id=? AND document_type=?", (next_number, company["id"], document_type))
            connection.commit(); flash(f"Imported {len(entries)} balanced transaction(s).", "success")
            return redirect(url_for("transactions"))
        except RuntimeError: flash("Excel import needs openpyxl. Run: python -m pip install -r requirements.txt", "error")
        except (ValueError, InvalidOperation, IndexError, sqlite3.IntegrityError):
            db().rollback(); flash("Import failed. Use unique document numbers, the required columns, and balanced entries.", "error")
    return render_template("import_transactions.html")


@app.route("/transactions/import/sample")
def import_sample():
    company = company_required()
    if not company: return redirect(url_for("companies_dashboard"))
    try:
        from openpyxl import Workbook
    except ModuleNotFoundError:
        return Response("<h2>Excel sample setup needed</h2><p>Please run: <code>python -m pip install -r requirements.txt</code>, then restart Ledgerly.</p>", status=503, mimetype="text/html")
    book = Workbook(); sheet = book.active; sheet.title = "Transactions"
    sheet.append(["Document Type", "Document No", "Date", "Reference", "Memo", "Account Ledger", "Description", "Debit", "Credit", "Currency", "FX Rate"])
    sheet.append(["Sales", suggested_document_no(company, "Sales"), display_date(company["financial_year_start"]), "INV-001", "Sample cash sale", "Bank account", "Bank received", 1000, "", company["base_currency"], 1])
    sheet.append(["Sales", suggested_document_no(company, "Sales"), display_date(company["financial_year_start"]), "INV-001", "Sample cash sale", "Sales revenue", "Sales revenue", "", 1000, company["base_currency"], 1])
    for column in sheet.columns: sheet.column_dimensions[column[0].column_letter].width = 18
    output = BytesIO(); book.save(output)
    return Response(output.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=ledgerly-transaction-import-sample.xlsx"})


@app.route("/accounts", methods=["GET", "POST"])
def accounts():
    company = company_required()
    if not company: return redirect(url_for("companies_dashboard"))
    if request.method == "POST":
        code, name, category = request.form.get("code", "").strip(), request.form.get("name", "").strip(), request.form.get("category", "")
        try:
            if not code or not name or category not in ("Asset", "Liability", "Equity", "Income", "Expense"): raise ValueError
            cash_type = request.form.get("cash_type", "Bank")
            if cash_type not in ("Cash", "Bank", "Card"): raise ValueError
            opening_date = normalise_date(request.form.get("opening_date", "") or company["financial_year_start"])
            opening_debit, opening_credit = Decimal(request.form.get("opening_debit") or 0), Decimal(request.form.get("opening_credit") or 0)
            if opening_debit < 0 or opening_credit < 0 or (opening_debit and opening_credit): raise ValueError
            default_currency = request.form.get("default_currency", company["base_currency"]).upper().strip()
            exchange_rate = Decimal(request.form.get("exchange_rate") or 1)
            subgroup_id = request.form.get("subgroup_id", type=int)
            subgroup = db().execute("SELECT 1 FROM account_subgroups WHERE id=? AND company_id=? AND category=?", (subgroup_id, company["id"], category)).fetchone() if subgroup_id else None
            if len(default_currency) != 3 or not default_currency.isalpha() or exchange_rate <= 0 or (subgroup_id and not subgroup): raise ValueError
            if default_currency == company["base_currency"]: exchange_rate = Decimal(1)
            connection = db(); account = connection.execute("INSERT INTO accounts(company_id,code,name,category,subgroup_id,is_cash,cash_type,is_loan,default_currency) VALUES(?,?,?,?,?,?,?,?,?)", (company["id"], code, name, category, subgroup_id, int(bool(request.form.get("is_cash"))), cash_type, int(category == "Liability" and bool(request.form.get("is_loan"))), default_currency))
            set_account_analysis_categories(company["id"], account.lastrowid, request.form.getlist("analysis_category_ids[]"))
            connection.execute("INSERT INTO currencies(company_id,code,name,rate_to_base) VALUES(?,?,?,?) ON CONFLICT(company_id,code) DO UPDATE SET rate_to_base=excluded.rate_to_base,active=1", (company["id"], default_currency, f"{default_currency} currency", float(exchange_rate)))
            if opening_debit or opening_credit:
                connection.execute("INSERT INTO opening_balances(company_id,account_id,effective_date,debit,credit,currency,fx_rate) VALUES(?,?,?,?,?,?,?)", (company["id"], account.lastrowid, opening_date, float(opening_debit), float(opening_credit), default_currency, float(exchange_rate)))
            connection.commit(); flash("Account created.", "success")
        except (ValueError, sqlite3.IntegrityError): flash("Use a unique account code and complete all fields.", "error")
        return redirect(url_for("accounts"))
    show_inactive = request.args.get("inactive") == "1"
    accounts_list = balances(company["id"])
    used_ids = {row[0] for row in db().execute("SELECT account_id FROM opening_balances WHERE company_id=? UNION SELECT l.account_id FROM journal_lines l JOIN journal_entries e ON e.id=l.entry_id WHERE e.company_id=?", (company["id"], company["id"])).fetchall()}
    if not show_inactive: accounts_list = [account for account in accounts_list if account["id"] in used_ids]
    totals = {"debit": sum(max(a["signed_balance"], 0) for a in accounts_list), "credit": sum(max(-a["signed_balance"], 0) for a in accounts_list)}
    group_totals = {}
    for group in ("Asset", "Liability", "Equity", "Income", "Expense"):
        group_totals[group] = {"debit": sum(max(a["signed_balance"], 0) for a in accounts_list if a["category"] == group), "credit": sum(max(-a["signed_balance"], 0) for a in accounts_list if a["category"] == group)}
    subgroups = db().execute("SELECT * FROM account_subgroups WHERE company_id=? ORDER BY category,name", (company["id"],)).fetchall()
    tags = db().execute("SELECT * FROM accounting_tags WHERE company_id=? AND active=1 ORDER BY name", (company["id"],)).fetchall()
    return render_template("accounts.html", accounts=accounts_list, totals=totals, group_totals=group_totals, show_inactive=show_inactive, subgroups=subgroups, tags=tags)


@app.route("/accounts/<int:account_id>/edit", methods=["GET", "POST"])
def edit_account(account_id):
    company = company_required()
    if not company: return redirect(url_for("companies_dashboard"))
    account = db().execute("SELECT * FROM accounts WHERE id=? AND company_id=?", (account_id, company["id"])).fetchone()
    if not account: return redirect(url_for("accounts"))
    opening = db().execute("SELECT * FROM opening_balances WHERE company_id=? AND account_id=? ORDER BY effective_date LIMIT 1", (company["id"], account_id)).fetchone()
    if request.method == "POST":
        try:
            code, name, category = request.form.get("code", "").strip(), request.form.get("name", "").strip(), request.form.get("category", "")
            debit, credit = Decimal(request.form.get("opening_debit") or 0), Decimal(request.form.get("opening_credit") or 0)
            opening_date = normalise_date(request.form.get("opening_date") or company["financial_year_start"])
            if not code or not name or category not in ("Asset", "Liability", "Equity", "Income", "Expense") or debit < 0 or credit < 0 or (debit and credit): raise ValueError
            default_currency = request.form.get("default_currency", company["base_currency"]).upper().strip()
            if len(default_currency) != 3 or not default_currency.isalpha(): raise ValueError
            subgroup_id = request.form.get("subgroup_id", type=int)
            subgroup = db().execute("SELECT 1 FROM account_subgroups WHERE id=? AND company_id=? AND category=?", (subgroup_id, company["id"], category)).fetchone() if subgroup_id else None
            if subgroup_id and not subgroup: raise ValueError
            exchange_rate = Decimal(request.form.get("exchange_rate") or (opening["fx_rate"] if opening else 1))
            if exchange_rate <= 0: raise ValueError
            if default_currency == company["base_currency"]: exchange_rate = Decimal(1)
            connection = db(); connection.execute("UPDATE accounts SET code=?,name=?,category=?,subgroup_id=?,is_cash=?,cash_type=?,is_loan=?,default_currency=? WHERE id=?", (code, name, category, subgroup_id, int(bool(request.form.get("is_cash"))), request.form.get("cash_type", "Bank"), int(category == "Liability" and bool(request.form.get("is_loan"))), default_currency, account_id))
            set_account_analysis_categories(company["id"], account_id, request.form.getlist("analysis_category_ids[]"))
            connection.execute("DELETE FROM opening_balances WHERE company_id=? AND account_id=?", (company["id"], account_id))
            if debit or credit: connection.execute("INSERT INTO opening_balances(company_id,account_id,effective_date,debit,credit,currency,fx_rate) VALUES(?,?,?,?,?,?,?)", (company["id"], account_id, opening_date, float(debit), float(credit), default_currency, float(exchange_rate)))
            connection.commit(); flash("Account updated.", "success"); return redirect(url_for("accounts"))
        except (ValueError, InvalidOperation, sqlite3.IntegrityError): flash("Check account fields and use only one opening balance side.", "error")
    subgroups = db().execute("SELECT * FROM account_subgroups WHERE company_id=? ORDER BY category,name", (company["id"],)).fetchall()
    tags = db().execute("SELECT * FROM accounting_tags WHERE company_id=? AND active=1 ORDER BY name", (company["id"],)).fetchall()
    tag_ids = {row[0] for row in db().execute("SELECT tag_id FROM account_tag_links WHERE account_id=?", (account_id,)).fetchall()}
    return render_template("edit_account.html", account=account, opening=opening, subgroups=subgroups, tags=tags, tag_ids=tag_ids)


@app.route("/accounts/<int:account_id>/delete", methods=["POST"])
def delete_account(account_id):
    company = company_required()
    if not company: return redirect(url_for("companies_dashboard"))
    try:
        db().execute("DELETE FROM accounts WHERE id=? AND company_id=?", (account_id, company["id"])); db().commit(); flash("Account deleted.", "success")
    except sqlite3.IntegrityError:
        db().rollback(); flash("This account has transactions and cannot be deleted. Keep it for the accounting history.", "error")
    return redirect(url_for("accounts", inactive="1"))


@app.route("/accounts/export")
def export_accounts():
    company = company_required()
    if not company: return redirect(url_for("companies_dashboard"))
    accounts_list = balances(company["id"]); rows = []
    for group in ("Asset", "Liability", "Equity", "Income", "Expense"):
        group_accounts = [a for a in accounts_list if a["category"] == group]
        for account in group_accounts:
            rows.append([group, account["code"], account["name"], money(max(account["signed_balance"], 0)), money(max(-account["signed_balance"], 0))])
        rows.append(["", "", f"TOTAL {group.upper()}S", money(sum(max(a["signed_balance"], 0) for a in group_accounts)), money(sum(max(-a["signed_balance"], 0) for a in group_accounts))])
    rows.append(["", "", "FINAL TOTAL", money(sum(max(a["signed_balance"], 0) for a in accounts_list)), money(sum(max(-a["signed_balance"], 0) for a in accounts_list))])
    return export_report(company, "Chart of Accounts", ["Group", "Code", "Account", "Debit", "Credit"], rows, company["base_currency"]) or redirect(url_for("accounts"))


@app.route("/accounts/import", methods=["GET", "POST"])
def import_accounts():
    company = company_required()
    if not company: return redirect(url_for("companies_dashboard"))
    if request.method == "POST":
        upload = request.files.get("file")
        try:
            if not upload or not upload.filename.lower().endswith(".xlsx"): raise ValueError
            try: from openpyxl import load_workbook
            except ModuleNotFoundError: raise RuntimeError
            sheet = load_workbook(upload, data_only=True).active
            headers = [str(c.value or "").strip().lower() for c in next(sheet.iter_rows(min_row=1, max_row=1))]
            if any(name not in headers for name in ("code", "name", "group")): raise ValueError
            index = {name: headers.index(name) for name in headers}; rows = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not any(row): continue
                code, name, group = str(row[index["code"]] or "").strip(), str(row[index["name"]] or "").strip(), str(row[index["group"]] or "").strip().title()
                cash = str(row[index["cash type"]] or "").strip().title() if "cash type" in index else ""
                if not code or not name or group not in ("Asset", "Liability", "Equity", "Income", "Expense") or cash not in ("", "Cash", "Bank", "Card"): raise ValueError
                opening_date = normalise_date(row[index["opening date"]]) if "opening date" in index and row[index["opening date"]] else company["financial_year_start"]
                opening_debit = Decimal(str(row[index["opening debit"]] or 0)) if "opening debit" in index else Decimal(0)
                opening_credit = Decimal(str(row[index["opening credit"]] or 0)) if "opening credit" in index else Decimal(0)
                if opening_debit < 0 or opening_credit < 0 or (opening_debit and opening_credit): raise ValueError
                default_currency = str(row[index["default currency"]] or company["base_currency"]).upper().strip() if "default currency" in index else company["base_currency"]
                if len(default_currency) != 3 or not default_currency.isalpha(): raise ValueError
                rows.append((code, name, group, cash, default_currency, opening_date, opening_debit, opening_credit))
            connection = db()
            for code, name, group, cash, default_currency, opening_date, opening_debit, opening_credit in rows:
                account = connection.execute("INSERT INTO accounts(company_id,code,name,category,is_cash,cash_type,default_currency) VALUES(?,?,?,?,?,?,?)", (company["id"], code, name, group, int(bool(cash)), cash or "Bank", default_currency))
                if opening_debit or opening_credit:
                    connection.execute("INSERT INTO opening_balances(company_id,account_id,effective_date,debit,credit) VALUES(?,?,?,?,?)", (company["id"], account.lastrowid, opening_date, float(opening_debit), float(opening_credit)))
            connection.commit()
            flash(f"Imported {len(rows)} account(s).", "success"); return redirect(url_for("accounts"))
        except RuntimeError: flash("Excel import needs openpyxl. Run: python -m pip install -r requirements.txt", "error")
        except (ValueError, sqlite3.IntegrityError, IndexError): flash("Import failed. Use unique codes and valid Code, Name, Group columns.", "error")
    return render_template("import_accounts.html")


@app.route("/accounts/import/sample")
def import_accounts_sample():
    company = company_required()
    if not company: return redirect(url_for("companies_dashboard"))
    try: from openpyxl import Workbook
    except ModuleNotFoundError: return Response("<h2>Excel sample setup needed</h2><p>Please run: <code>python -m pip install -r requirements.txt</code>.</p>", status=503, mimetype="text/html")
    book = Workbook(); sheet = book.active; sheet.title = "Chart of Accounts"
    sheet.append(["Code", "Name", "Group", "Cash Type", "Default Currency", "Opening Date", "Opening Debit", "Opening Credit"])
    sheet.append(["1010", "Petty Cash", "Asset", "Cash", company["base_currency"], display_date(company["financial_year_start"]), 500, ""]); sheet.append(["6300", "Travel Expense", "Expense", "", company["base_currency"], display_date(company["financial_year_start"]), "", 0])
    for column in sheet.columns: sheet.column_dimensions[column[0].column_letter].width = 22
    output = BytesIO(); book.save(output)
    return Response(output.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=ledgerly-chart-of-accounts-sample.xlsx"})


@app.route("/banking")
def banking():
    company = company_required()
    if not company: return redirect(url_for("companies_dashboard"))
    cash_accounts = db().execute("SELECT * FROM accounts WHERE company_id=? AND is_cash=1 ORDER BY code", (company["id"],)).fetchall()
    account_id = request.args.get("account_id", type=int)
    all_balances = {row["id"]: row["signed_balance"] for row in balances(company["id"])}
    currency_rates = {row["code"]: row["rate_to_base"] for row in db().execute("SELECT code,rate_to_base FROM currencies WHERE company_id=?", (company["id"],)).fetchall()}
    bank_summary = []
    for account in cash_accounts:
        statement = db().execute("SELECT COALESCE(SUM(debit-credit),0) total,COUNT(*) lines,COALESCE(SUM(CASE WHEN matched_journal_line_id IS NULL THEN 1 ELSE 0 END),0) unmatched FROM bank_statement_lines WHERE company_id=? AND account_id=?", (company["id"], account["id"])).fetchone()
        statement_opening = db().execute("SELECT COALESCE(SUM(debit-credit),0) amount FROM bank_statement_openings WHERE company_id=? AND account_id=?", (company["id"], account["id"])).fetchone()["amount"]
        rate = currency_rates.get(account["default_currency"], 1) or 1
        account_amount = account_balance_in_currency(company["id"], account["id"], account["default_currency"], rate)
        statement_amount = statement["total"] + statement_opening
        bank_summary.append({**dict(account), "statement_balance": statement_amount, "statement_lines": statement["lines"], "unmatched": statement["unmatched"], "account_balance": account_amount, "status": "Reconciled" if not statement["unmatched"] and abs(statement_amount-account_amount) < .01 else "Unreconciled"})
    selected_account = next((row for row in cash_accounts if row["id"] == account_id), None)
    if account_id and not selected_account: account_id = None
    statement_rows, statement_display, candidates, ledger, account_openings = [], [], {}, [], []
    account_currency = selected_account["default_currency"] if selected_account else company["base_currency"]
    if account_id:
        statement_rows = db().execute("SELECT * FROM bank_statement_lines WHERE company_id=? AND account_id=? ORDER BY statement_date DESC,id DESC", (company["id"], account_id)).fetchall()
        statement_opening = db().execute("SELECT * FROM bank_statement_openings WHERE company_id=? AND account_id=?", (company["id"], account_id)).fetchone()
        if statement_opening:
            statement_display.append({**dict(statement_opening), "description": "Opening balance", "reference": "Opening balance", "is_opening": True, "matched_journal_line_id": 1})
        statement_display.extend([{**dict(row), "is_opening": False} for row in sorted(statement_rows, key=lambda row: (row["statement_date"], row["id"]))])
        statement_running = 0
        for row in statement_display:
            statement_running += row["debit"] - row["credit"]
            row["running_balance"] = statement_running
        ledger_rows = db().execute("""SELECT l.id,e.entry_date,e.document_no,e.reference,e.memo,l.description,l.debit,l.credit,l.currency,l.fx_rate,
          CASE WHEN l.id IN (SELECT COALESCE(matched_journal_line_id,-1) FROM bank_statement_lines WHERE company_id=?) THEN 1 ELSE 0 END reconciled
          FROM journal_lines l JOIN journal_entries e ON e.id=l.entry_id WHERE e.company_id=? AND l.account_id=? ORDER BY e.entry_date DESC,l.id DESC""", (company["id"], company["id"], account_id)).fetchall()
        account_rate = currency_rates.get(account_currency, 1) or 1
        ledger = [{**dict(line), "debit": line["debit"] if line["currency"] == account_currency else line["debit"] * line["fx_rate"] / account_rate, "credit": line["credit"] if line["currency"] == account_currency else line["credit"] * line["fx_rate"] / account_rate} for line in ledger_rows]
        account_openings = db().execute("SELECT * FROM opening_balances WHERE company_id=? AND account_id=? ORDER BY effective_date", (company["id"], account_id)).fetchall()
        ledger = [{"id": None, "entry_date": row["effective_date"], "document_no": "OPENING", "reference": "", "memo": "", "description": "Opening balance", "debit": row["debit"] * row["fx_rate"] / account_rate, "credit": row["credit"] * row["fx_rate"] / account_rate, "reconciled": 1, "is_opening": True} for row in account_openings] + [{**line, "is_opening": False} for line in ledger]
        ledger.sort(key=lambda row: (row["entry_date"], 0 if row["is_opening"] else 1, row["id"] or 0))
        ledger_running = 0
        for line in ledger:
            ledger_running += line["debit"] - line["credit"]
            line["running_balance"] = ledger_running
        for row in statement_rows:
            candidates[row["id"]] = [line for line in ledger if not line["reconciled"] and abs((line["debit"]-line["credit"]) - (row["debit"]-row["credit"])) < .01]
    statement_balance = sum(row["debit"]-row["credit"] for row in statement_rows) + (statement_opening["debit"]-statement_opening["credit"] if account_id and statement_opening else 0)
    account_rate = currency_rates.get(account_currency, 1) or 1
    account_balance = account_balance_in_currency(company["id"], account_id, account_currency, account_rate) if account_id else 0
    counterparts = db().execute("SELECT * FROM accounts WHERE company_id=? AND active=1 AND id<>? ORDER BY code", (company["id"], account_id or -1)).fetchall()
    unreconciled_rows = [row for row in statement_rows if not row["matched_journal_line_id"]]
    return render_template("banking.html", cash_accounts=cash_accounts, bank_summary=bank_summary, account_id=account_id, selected_account=selected_account, account_currency=account_currency, statement_rows=statement_rows, unreconciled_rows=unreconciled_rows, statement_display=statement_display, statement_opening=statement_opening if account_id else None, edit_statement_opening=request.args.get("edit_opening") == "1", candidates=candidates, ledger=ledger, statement_balance=statement_balance, account_balance=account_balance, counterparts=counterparts)


@app.route("/accounts/quick-add", methods=["POST"])
def quick_add_account():
    company = company_required()
    if not company: return {"error": "Select a company first"}, 400
    try:
        code, name, category = request.form.get("code", "").strip(), request.form.get("name", "").strip(), request.form.get("category", "")
        if not code or not name or category not in ("Asset", "Liability", "Equity", "Income", "Expense"): raise ValueError
        account = db().execute("INSERT INTO accounts(company_id,code,name,category,is_cash,cash_type,default_currency) VALUES(?,?,?,?,0,'Bank',?)", (company["id"], code, name, category, company["base_currency"]))
        db().commit()
        return {"id": account.lastrowid, "code": code, "name": name}
    except (ValueError, sqlite3.IntegrityError):
        db().rollback(); return {"error": "Use a unique account code, name, and valid group."}, 400


@app.route("/banking/statement-opening", methods=["POST"])
def save_bank_statement_opening():
    company = company_required()
    if not company: return redirect(url_for("companies_dashboard"))
    try:
        account_id = int(request.form["account_id"])
        effective_date = normalise_date(request.form["effective_date"])
        debit, credit = Decimal(request.form.get("debit") or 0), Decimal(request.form.get("credit") or 0)
        valid = db().execute("SELECT 1 FROM accounts WHERE id=? AND company_id=? AND is_cash=1", (account_id, company["id"])).fetchone()
        if not valid or debit < 0 or credit < 0 or (debit and credit): raise ValueError
        db().execute("INSERT INTO bank_statement_openings(company_id,account_id,effective_date,debit,credit) VALUES(?,?,?,?,?) ON CONFLICT(company_id,account_id) DO UPDATE SET effective_date=excluded.effective_date,debit=excluded.debit,credit=excluded.credit", (company["id"], account_id, effective_date, float(debit), float(credit)))
        db().commit(); flash("Bank statement opening balance saved.", "success")
    except (KeyError, ValueError, InvalidOperation):
        flash("Enter one valid debit or credit opening balance.", "error")
    return redirect(url_for("banking", account_id=request.form.get("account_id")))


@app.route("/banking/statement/<int:statement_id>/edit", methods=["GET", "POST"])
def edit_bank_statement_line(statement_id):
    company = company_required()
    if not company: return redirect(url_for("companies_dashboard"))
    line = db().execute("SELECT s.*,a.code,a.name,a.default_currency FROM bank_statement_lines s JOIN accounts a ON a.id=s.account_id WHERE s.id=? AND s.company_id=?", (statement_id, company["id"])).fetchone()
    if not line: return redirect(url_for("banking"))
    if request.method == "POST":
        try:
            statement_date = normalise_date(request.form["statement_date"])
            debit, credit = Decimal(request.form.get("debit") or 0), Decimal(request.form.get("credit") or 0)
            if debit < 0 or credit < 0 or (debit and credit): raise ValueError
            db().execute("UPDATE bank_statement_lines SET statement_date=?,description=?,reference=?,debit=?,credit=?,matched_journal_line_id=NULL WHERE id=?", (statement_date, request.form.get("description", "").strip(), request.form.get("reference", "").strip(), float(debit), float(credit), statement_id))
            db().commit(); flash("Bank statement line updated. Please reconcile it again if required.", "success")
            return redirect(url_for("banking", account_id=line["account_id"]) + "#statement")
        except (KeyError, ValueError, InvalidOperation):
            flash("Enter a valid date and one deposit or withdrawal amount.", "error")
    return render_template("edit_bank_statement.html", line=line)


@app.route("/banking/statement/<int:statement_id>/delete", methods=["POST"])
def delete_bank_statement_line(statement_id):
    company = company_required()
    if not company: return redirect(url_for("companies_dashboard"))
    line = db().execute("SELECT account_id FROM bank_statement_lines WHERE id=? AND company_id=?", (statement_id, company["id"])).fetchone()
    if line:
        db().execute("DELETE FROM bank_statement_lines WHERE id=? AND company_id=?", (statement_id, company["id"])); db().commit(); flash("Bank statement line deleted.", "success")
    return redirect(url_for("banking", account_id=line["account_id"] if line else None) + "#statement")


@app.route("/banking/statement/delete-selected", methods=["POST"])
def delete_selected_bank_statement_lines():
    company = company_required()
    if not company: return redirect(url_for("companies_dashboard"))
    account_id = request.form.get("account_id", type=int)
    selected = [int(item) for item in request.form.getlist("statement_ids") if item.isdigit()]
    if selected:
        marks = ",".join("?" for _ in selected)
        db().execute(f"DELETE FROM bank_statement_lines WHERE company_id=? AND account_id=? AND id IN ({marks})", [company["id"], account_id, *selected]); db().commit(); flash(f"Deleted {len(selected)} bank statement line(s).", "success")
    return redirect(url_for("banking", account_id=account_id) + "#statement")


@app.route("/banking/create-and-reconcile", methods=["POST"])
def create_and_reconcile():
    company = company_required()
    if not company: return redirect(url_for("companies_dashboard"))
    account_id = request.form.get("account_id", type=int); counterpart_id = request.form.get("counterpart_id", type=int)
    selected = [int(value) for value in request.form.getlist("statement_ids") if value.isdigit()]
    valid = db().execute("SELECT 1 FROM accounts WHERE id=? AND company_id=?", (counterpart_id, company["id"])).fetchone()
    rows = db().execute(f"SELECT * FROM bank_statement_lines WHERE company_id=? AND account_id=? AND matched_journal_line_id IS NULL AND id IN ({','.join('?' for _ in selected) or 'NULL'})", [company["id"], account_id, *selected]).fetchall()
    if not rows or not valid: flash("Select unmatched statement lines and a counterpart account.", "error")
    else:
        connection=db()
        for row in rows:
            entry=connection.execute("INSERT INTO journal_entries(company_id,entry_date,document_type,document_no,reference,memo) VALUES(?,?,?,?,?,?)",(company["id"],row["statement_date"],"Journal",f"BANK-{row['id']}",row["reference"],row["description"]))
            bank_debit,bank_credit=row["debit"],row["credit"]
            bank_line=connection.execute("INSERT INTO journal_lines(entry_id,account_id,description,debit,credit,currency,fx_rate) VALUES(?,?,?,?,?,?,1)",(entry.lastrowid,account_id,row["description"],bank_debit,bank_credit,company["base_currency"]))
            connection.execute("INSERT INTO journal_lines(entry_id,account_id,description,debit,credit,currency,fx_rate) VALUES(?,?,?,?,?,?,1)",(entry.lastrowid,counterpart_id,row["description"],bank_credit,bank_debit,company["base_currency"]))
            connection.execute("UPDATE bank_statement_lines SET matched_journal_line_id=? WHERE id=?",(bank_line.lastrowid,row["id"]))
        connection.commit(); flash(f"Created and reconciled {len(rows)} transaction(s).","success")
    return redirect(url_for("banking",account_id=account_id))


@app.route("/banking/statement/<int:statement_id>/create-transaction", methods=["GET", "POST"])
def create_transaction_from_statement(statement_id):
    company = company_required()
    if not company: return redirect(url_for("companies_dashboard"))
    statement = db().execute("""SELECT s.*,a.code bank_code,a.name bank_name,a.default_currency
      FROM bank_statement_lines s JOIN accounts a ON a.id=s.account_id
      WHERE s.id=? AND s.company_id=? AND s.matched_journal_line_id IS NULL""", (statement_id, company["id"])).fetchone()
    if not statement:
        flash("This statement line is already reconciled or unavailable.", "error")
        return redirect(url_for("banking"))
    accounts = db().execute("SELECT * FROM accounts WHERE company_id=? AND active=1 AND id<>? ORDER BY code", (company["id"], statement["account_id"])).fetchall()
    currency_rates = {row["code"]: row["rate_to_base"] for row in db().execute("SELECT code,rate_to_base FROM currencies WHERE company_id=?", (company["id"],)).fetchall()}
    bank_rate = currency_rates.get(statement["default_currency"], 1) or 1
    if request.method == "POST":
        try:
            entry_date = normalise_date(request.form["entry_date"])
            document_type = request.form.get("document_type", "Journal")
            if document_type not in DOCUMENT_TYPES or not company["financial_year_start"] <= entry_date <= company["financial_year_end"]: raise ValueError
            series = document_series(company["id"], document_type)
            document_no = suggested_document_no(company, document_type) if series["number_mode"] == "automatic" else request.form.get("document_no", "").strip()
            if not document_no: raise ValueError
            lines = [(statement["account_id"], statement["description"] or "Bank statement", float(statement["debit"]), float(statement["credit"]), statement["default_currency"], float(bank_rate))]
            for account, description, debit, credit, currency, rate in zip(request.form.getlist("account_id[]"), request.form.getlist("description[]"), request.form.getlist("debit[]"), request.form.getlist("credit[]"), request.form.getlist("currency[]"), request.form.getlist("fx_rate[]")):
                debit, credit, rate = Decimal(debit or 0), Decimal(credit or 0), Decimal(rate or 1)
                if int(account) not in [row["id"] for row in accounts] or debit < 0 or credit < 0 or (debit and credit) or rate <= 0: raise ValueError
                if debit or credit: lines.append((int(account), description.strip(), float(debit), float(credit), currency.upper().strip(), float(rate)))
            # If both sides use the same currency, the rate entered for the
            # allocation is the transaction's agreed rate. Preserve it on the
            # bank line as well rather than replacing it with the master rate.
            if len(lines) == 2 and lines[0][4] == lines[1][4]:
                bank = list(lines[0]); bank[5] = lines[1][5]; lines[0] = tuple(bank)
            if len(lines) < 2 or abs(sum(Decimal(str(line[2])) * Decimal(str(line[5])) for line in lines) - sum(Decimal(str(line[3])) * Decimal(str(line[5])) for line in lines)) > Decimal(".005"): raise ValueError
            payment_mode = request.form.get("payment_mode", "").strip() if document_type in ("Payments", "Receipts") else ""
            if payment_mode not in ("", "Cash", "Bank", "Card", "Other"): raise ValueError
            tag_value = next((value for value in request.form.getlist("accounting_tag_id[]") if value), "")
            tag_id, region_id = (int(tag_value) if tag_value.isdigit() else None), request.form.get("region_tag_id", type=int)
            connection = db(); entry = connection.execute("INSERT INTO journal_entries(company_id,entry_date,document_type,document_no,reference,memo,payment_mode,party,accounting_tag_id,region_tag_id) VALUES(?,?,?,?,?,?,?,?,?,?)", (company["id"], entry_date, document_type, document_no, request.form.get("reference", "").strip(), request.form.get("memo", "").strip(), payment_mode, request.form.get("party", "").strip(), tag_id, region_id))
            bank_line = connection.execute("INSERT INTO journal_lines(entry_id,account_id,description,debit,credit,currency,fx_rate) VALUES(?,?,?,?,?,?,?)", (entry.lastrowid, *lines[0]))
            connection.executemany("INSERT INTO journal_lines(entry_id,account_id,description,debit,credit,currency,fx_rate) VALUES(?,?,?,?,?,?,?)", [(entry.lastrowid, *line) for line in lines[1:]])
            connection.execute("UPDATE bank_statement_lines SET matched_journal_line_id=? WHERE id=?", (bank_line.lastrowid, statement_id))
            if series["number_mode"] == "automatic": connection.execute("UPDATE document_series SET next_number=next_number+1 WHERE company_id=? AND document_type=?", (company["id"], document_type))
            connection.commit(); flash("Transaction posted and bank statement reconciled.", "success")
            return redirect(url_for("banking", account_id=statement["account_id"]) + "#accounting")
        except (KeyError, ValueError, InvalidOperation, sqlite3.IntegrityError):
            db().rollback(); flash("Add one or more ledger lines and ensure deposits equal withdrawals.", "error")
    return render_template("create_bank_transaction.html", statement=statement, accounts=accounts, document_types=DOCUMENT_TYPES, suggested_document_no=suggested_document_no(company), bank_rate=bank_rate)


@app.route("/banking/import", methods=["GET", "POST"])
def import_bank_statement():
    company = company_required()
    if not company: return redirect(url_for("companies_dashboard"))
    accounts = db().execute("SELECT * FROM accounts WHERE company_id=? AND is_cash=1 ORDER BY code", (company["id"],)).fetchall()
    if request.method == "POST":
        upload, account_id = request.files.get("file"), request.form.get("account_id", type=int)
        try:
            if not upload or not upload.filename.lower().endswith(".xlsx") or not db().execute("SELECT 1 FROM accounts WHERE id=? AND company_id=? AND is_cash=1", (account_id, company["id"])).fetchone(): raise ValueError
            try: from openpyxl import load_workbook
            except ModuleNotFoundError: raise RuntimeError
            sheet = load_workbook(upload, data_only=True).active; headers = [str(c.value or "").strip().lower() for c in next(sheet.iter_rows(min_row=1,max_row=1))]
            if any(name not in headers for name in ("date", "description")): raise ValueError
            deposit_column = "deposits" if "deposits" in headers else "debit"
            withdrawal_column = "withdrawals" if "withdrawals" in headers else "credit"
            if deposit_column not in headers or withdrawal_column not in headers: raise ValueError
            index = {name: headers.index(name) for name in headers}; lines=[]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not any(row): continue
                statement_date = normalise_date(row[index["date"]])
                debit, credit = Decimal(str(row[index[deposit_column]] or 0)), Decimal(str(row[index[withdrawal_column]] or 0))
                if debit < 0 or credit < 0 or (debit and credit): raise ValueError
                lines.append((company["id"], account_id, statement_date, str(row[index["description"]] or ""), str(row[index.get("reference",-1)] or "") if "reference" in index else "", float(debit), float(credit)))
            db().executemany("INSERT INTO bank_statement_lines(company_id,account_id,statement_date,description,reference,debit,credit) VALUES(?,?,?,?,?,?,?)", lines); db().commit(); flash(f"Imported {len(lines)} statement line(s).", "success")
            return redirect(url_for("banking", account_id=account_id))
        except RuntimeError: flash("Statement import needs openpyxl. Run: python -m pip install -r requirements.txt", "error")
        except (ValueError, InvalidOperation, IndexError): flash("Import failed. Use Date, Description, Deposits and Withdrawals columns.", "error")
    return render_template("import_bank_statement.html", accounts=accounts)


@app.route("/banking/import/sample")
def import_bank_statement_sample():
    company = company_required()
    if not company: return redirect(url_for("companies_dashboard"))
    try: from openpyxl import Workbook
    except ModuleNotFoundError: return Response("<h2>Excel sample setup needed</h2><p>Please run: <code>python -m pip install -r requirements.txt</code>.</p>", status=503, mimetype="text/html")
    book = Workbook(); sheet = book.active; sheet.title = "Bank Statement"
    sheet.append(["Date", "Description", "Reference", "Deposits", "Withdrawals"])
    sheet.append([display_date(company["financial_year_start"]), "Customer payment received", "BANK-001", 2500, ""])
    sheet.append([display_date(company["financial_year_start"]), "Bank charges", "BANK-002", "", 25])
    for column in sheet.columns: sheet.column_dimensions[column[0].column_letter].width = 24
    output = BytesIO(); book.save(output)
    return Response(output.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=ledgerly-bank-statement-import-sample.xlsx"})


@app.route("/banking/reconcile", methods=["POST"])
def reconcile_statement():
    company = company_required()
    if not company: return redirect(url_for("companies_dashboard"))
    statement_id, line_id = request.form.get("statement_id", type=int), request.form.get("journal_line_id", type=int)
    statement = db().execute("SELECT * FROM bank_statement_lines WHERE id=? AND company_id=?", (statement_id, company["id"])).fetchone()
    valid_line = db().execute("SELECT l.id FROM journal_lines l JOIN journal_entries e ON e.id=l.entry_id WHERE l.id=? AND e.company_id=? AND l.account_id=?", (line_id, company["id"], statement["account_id"] if statement else -1)).fetchone()
    if statement and valid_line:
        db().execute("UPDATE bank_statement_lines SET matched_journal_line_id=? WHERE id=?", (line_id, statement_id)); db().commit(); flash("Statement line reconciled with the accounting transaction.", "success")
    return redirect(url_for("banking", account_id=statement["account_id"] if statement else None))


@app.route("/reports")
def reports_home():
    company = company_required()
    if not company: return redirect(url_for("companies_dashboard"))
    return render_template("reports_menu.html")


@app.route("/accounts-book/transaction-register")
@app.route("/reports/document-types")
def document_type_report():
    company = company_required()
    if not company: return redirect(url_for("companies_dashboard"))
    start, end = report_period(company); currency_code, factor = report_factor(company)
    selected_type = request.args.get("document_type", "")
    if selected_type and selected_type not in DOCUMENT_TYPES: selected_type = ""
    summary_rows = db().execute("""SELECT e.document_type,COUNT(DISTINCT e.id) transaction_count,COALESCE(SUM(l.debit*l.fx_rate),0) debit,COALESCE(SUM(l.credit*l.fx_rate),0) credit FROM journal_entries e LEFT JOIN journal_lines l ON l.entry_id=e.id WHERE e.company_id=? AND e.entry_date BETWEEN ? AND ? GROUP BY e.document_type""", (company["id"], start, end)).fetchall()
    summary = {row["document_type"]: dict(row) for row in summary_rows}
    summary_rows = [{"document_type": kind, "transaction_count": summary.get(kind, {}).get("transaction_count", 0), "debit": summary.get(kind, {}).get("debit", 0), "credit": summary.get(kind, {}).get("credit", 0)} for kind in DOCUMENT_TYPES]
    sql = """SELECT e.id,e.entry_date,e.document_type,e.document_no,e.reference,e.memo,COALESCE(SUM(l.debit*l.fx_rate),0) debit,COALESCE(SUM(l.credit*l.fx_rate),0) credit FROM journal_entries e LEFT JOIN journal_lines l ON l.entry_id=e.id WHERE e.company_id=? AND e.entry_date BETWEEN ? AND ?"""; params = [company["id"], start, end]
    if selected_type: sql += " AND e.document_type=?"; params.append(selected_type)
    detail_rows = db().execute(sql + " GROUP BY e.id ORDER BY e.entry_date,e.id", params).fetchall()
    if factor != 1:
        summary_rows = [{**row, "debit": row["debit"] * float(factor), "credit": row["credit"] * float(factor)} for row in summary_rows]
        detail_rows = [{**dict(row), "debit": row["debit"] * float(factor), "credit": row["credit"] * float(factor)} for row in detail_rows]
    export_rows = [[display_date(r["entry_date"]), r["document_type"], r["document_no"], r["reference"] or "", r["memo"] or "", money(r["debit"]), money(r["credit"])] for r in detail_rows]
    result = export_report(company, "Transaction Register", ["Date", "Document Type", "Document No.", "Reference", "Narration", "Debit", "Credit"], export_rows, currency_code)
    if result: return result
    return render_template("document_type_report.html", summary_rows=summary_rows, detail_rows=detail_rows, selected_type=selected_type, start=start, end=end, base_currency=currency_code, report_rate=str(factor))


@app.route("/reports/analysis-tags")
def analysis_tag_report():
    company = company_required()
    if not company: return redirect(url_for("companies_dashboard"))
    start, end = report_period(company); currency_code, factor = report_factor(company); category_ids = [int(value) for value in request.args.getlist("category_ids") if value.isdigit()]
    sql = """SELECT e.id,e.entry_date,e.document_type,e.document_no,e.reference,e.memo,e.party,c.name category,t.name tag,COALESCE(SUM(l.debit*l.fx_rate),0) debit,COALESCE(SUM(l.credit*l.fx_rate),0) credit FROM journal_entries e JOIN accounting_tags t ON t.id=e.accounting_tag_id JOIN analysis_categories c ON c.id=t.analysis_category_id JOIN journal_lines l ON l.entry_id=e.id WHERE e.company_id=? AND e.entry_date BETWEEN ? AND ?"""; params = [company["id"], start, end]
    if category_ids: sql += " AND c.id IN (%s)" % ",".join("?" * len(category_ids)); params.extend(category_ids)
    rows = db().execute(sql + " GROUP BY e.id ORDER BY c.name,t.name,e.entry_date,e.id", params).fetchall()
    if factor != 1: rows = [{**dict(row), "debit": row["debit"] * float(factor), "credit": row["credit"] * float(factor)} for row in rows]
    export_rows = [[display_date(r["entry_date"]), r["category"], r["tag"], r["document_type"], r["document_no"], r["party"] or "", r["reference"] or "", r["memo"] or "", money(r["debit"]), money(r["credit"])] for r in rows]
    result = export_report(company, "Analysis Tag Transactions", ["Date", "Analysis Category", "Analysis Tag", "Document Type", "Document No.", "Party", "Reference", "Narration", "Debit", "Credit"], export_rows, currency_code)
    if result: return result
    categories = db().execute("SELECT * FROM analysis_categories WHERE company_id=? ORDER BY name", (company["id"],)).fetchall()
    return render_template("analysis_tag_report.html", rows=rows, categories=categories, category_ids=category_ids, start=start, end=end, base_currency=currency_code, report_rate=str(factor))


@app.route("/accounts-book/transaction-register/<document_type>")
@app.route("/reports/document-types/<document_type>")
def document_type_detail(document_type):
    if document_type not in DOCUMENT_TYPES: return redirect(url_for("document_type_report"))
    return redirect(url_for("document_type_report", document_type=document_type, period_from=request.args.get("period_from", ""), period_to=request.args.get("period_to", "")))


@app.route("/reports/day-book")
def day_book():
    company = company_required()
    if not company: return redirect(url_for("companies_dashboard"))
    today = date.today().isoformat()
    start = request.args.get("period_from") or request.args.get("from") or today
    end = request.args.get("period_to") or request.args.get("to") or today
    try:
        if datetime.strptime(start, "%Y-%m-%d") > datetime.strptime(end, "%Y-%m-%d"): raise ValueError
    except ValueError: start = end = today
    currency_code, factor = report_factor(company)
    rows = db().execute("""SELECT e.id,e.entry_date,e.document_type,e.document_no,e.reference,e.memo,COALESCE(SUM(l.debit*l.fx_rate),0) debit,COALESCE(SUM(l.credit*l.fx_rate),0) credit FROM journal_entries e LEFT JOIN journal_lines l ON l.entry_id=e.id WHERE e.company_id=? AND e.entry_date BETWEEN ? AND ? GROUP BY e.id ORDER BY e.entry_date,e.id""", (company["id"], start, end)).fetchall()
    if factor != 1: rows = [{**dict(row), "debit": row["debit"] * float(factor), "credit": row["credit"] * float(factor)} for row in rows]
    export_rows = [[display_date(r["entry_date"]), r["document_type"], r["document_no"], r["reference"] or "", r["memo"] or "", money(r["debit"]), money(r["credit"])] for r in rows]
    export_rows.append(["TOTAL", "", "", "", "", money(sum(r["debit"] for r in rows)), money(sum(r["credit"] for r in rows))])
    result = export_report(company, "Day Book", ["Date", "Document Type", "Document No.", "Reference", "Memo", "Debit", "Credit"], export_rows, currency_code)
    if result: return result
    return render_template("day_book.html", rows=rows, start=start, end=end, total_debit=sum(r["debit"] for r in rows), total_credit=sum(r["credit"] for r in rows), base_currency=currency_code, report_rate=str(factor))


@app.route("/reports/trial-balance")
def trial_balance():
    company = company_required()
    if not company: return redirect(url_for("companies_dashboard"))
    start, end = report_period(company); currency_code, factor = report_factor(company)
    opening_date = (date.fromisoformat(start) - timedelta(days=1)).isoformat()
    opening = balances(company["id"], opening_date)
    # A balance entered through the Opening Balances screen is always an opening
    # figure. Show every saved opening up to the selected report end date in the
    # opening columns, regardless of the effective date selected when entering it.
    saved_openings = opening_adjustments(company["id"], end)
    prior_openings = opening_adjustments(company["id"], opening_date)
    opening = [{**account, "signed_balance": account["signed_balance"] + saved_openings.get(account["id"], 0) - prior_openings.get(account["id"], 0)} for account in opening]
    period = movements(company["id"], start, end)
    rows = []
    for account in opening:
        opening_signed = account["signed_balance"]
        debit, credit = period.get(account["id"], (0, 0)); closing = opening_signed + debit - credit
        rows.append({**account, "opening_debit": max(opening_signed, 0), "opening_credit": max(-opening_signed, 0), "period_debit": debit, "period_credit": credit, "closing_debit": max(closing, 0), "closing_credit": max(-closing, 0)})
    totals = {key: sum(row[key] for row in rows) for key in ("opening_debit", "opening_credit", "period_debit", "period_credit", "closing_debit", "closing_credit")}
    opening_difference = totals["opening_debit"] - totals["opening_credit"]
    period_difference = totals["period_debit"] - totals["period_credit"]
    closing_difference = totals["closing_debit"] - totals["closing_credit"]
    if any(abs(value) > .005 for value in (opening_difference, period_difference, closing_difference)):
        rows.append({"id": None, "code": "", "name": "Diff in opening balance", "is_difference": True, "opening_debit": max(-opening_difference, 0), "opening_credit": max(opening_difference, 0), "period_debit": max(-period_difference, 0), "period_credit": max(period_difference, 0), "closing_debit": max(-closing_difference, 0), "closing_credit": max(closing_difference, 0)})
        totals = {key: sum(row[key] for row in rows) for key in ("opening_debit", "opening_credit", "period_debit", "period_credit", "closing_debit", "closing_credit")}
    if factor != 1:
        for row in rows:
            for key in ("opening_debit", "opening_credit", "period_debit", "period_credit", "closing_debit", "closing_credit"): row[key] *= float(factor)
        totals = {key: value * float(factor) for key, value in totals.items()}
    show_opening = request.args.get("opening") == "1"
    show_period = request.args.get("period") == "1"
    headers = ["Account"]
    if show_opening: headers += ["Opening Debit", "Opening Credit"]
    if show_period: headers += ["Period Debit", "Period Credit"]
    headers += ["Closing Debit", "Closing Credit"]
    export_rows = []
    for row in rows:
        values = [row["name"]]
        if show_opening: values += [money(row["opening_debit"]), money(row["opening_credit"])]
        if show_period: values += [money(row["period_debit"]), money(row["period_credit"])]
        values += [money(row["closing_debit"]), money(row["closing_credit"])]
        export_rows.append(values)
    total_row = ["TOTAL"]
    if show_opening: total_row += [money(totals["opening_debit"]), money(totals["opening_credit"])]
    if show_period: total_row += [money(totals["period_debit"]), money(totals["period_credit"])]
    total_row += [money(totals["closing_debit"]), money(totals["closing_credit"])]
    export_rows.append(total_row)
    result = export_report(company, "Trial Balance", headers, export_rows, currency_code)
    if result: return result
    return render_template("trial_balance.html", rows=rows, totals=totals, start=start, end=end, show_opening=show_opening, show_period=show_period, base_currency=currency_code, report_rate=str(factor))


@app.route("/reports/opening-balances", methods=["GET", "POST"])
def opening_balances():
    company = company_required()
    if not company: return redirect(url_for("companies_dashboard"))
    accounts = db().execute("SELECT * FROM accounts WHERE company_id=? ORDER BY code", (company["id"],)).fetchall()
    if request.method == "POST":
        try:
            account_id = int(request.form["account_id"]); effective_date = normalise_date(request.form["effective_date"])
            debit, credit = Decimal(request.form.get("debit") or 0), Decimal(request.form.get("credit") or 0)
            account = db().execute("SELECT default_currency,is_loan FROM accounts WHERE id=? AND company_id=?", (account_id, company["id"])).fetchone()
            if debit < 0 or credit < 0 or (debit and credit) or not account: raise ValueError
            currency = account["default_currency"] or company["base_currency"]
            rate_row = db().execute("SELECT rate_to_base FROM currencies WHERE company_id=? AND code=?", (company["id"], currency)).fetchone()
            rate = 1 if currency == company["base_currency"] else (rate_row["rate_to_base"] if rate_row else 1)
            db().execute("INSERT INTO opening_balances(company_id,account_id,effective_date,debit,credit,currency,fx_rate) VALUES(?,?,?,?,?,?,?) ON CONFLICT(company_id,account_id,effective_date) DO UPDATE SET debit=excluded.debit,credit=excluded.credit,currency=excluded.currency,fx_rate=excluded.fx_rate", (company["id"], account_id, effective_date, float(debit), float(credit), currency, rate))
            paid_count = allocate_loan_opening_schedule(company["id"], account_id) if account["is_loan"] else 0
            db().commit(); flash(f"Opening balance saved.{f' {paid_count} repayment schedule line(s) marked paid.' if paid_count else ''}", "success")
        except (KeyError, ValueError, InvalidOperation): flash("Enter one positive debit or credit and select an account.", "error")
        return redirect(url_for("opening_balances"))
    saved_raw = db().execute("SELECT o.*,a.code,a.name FROM opening_balances o JOIN accounts a ON a.id=o.account_id WHERE o.company_id=? ORDER BY o.effective_date,a.code", (company["id"],)).fetchall()
    saved = [{**dict(row), "display_debit": float(row["debit"] or 0) * float(row["fx_rate"] or 1), "display_credit": float(row["credit"] or 0) * float(row["fx_rate"] or 1)} for row in saved_raw]
    opening_totals = {"debit": sum(row["display_debit"] for row in saved), "credit": sum(row["display_credit"] for row in saved)}
    opening_difference = opening_totals["debit"] - opening_totals["credit"]
    return render_template("opening_balances.html", accounts=accounts, saved=saved, opening_totals=opening_totals, opening_difference=opening_difference)


@app.route("/reports/income-statement")
def income_statement():
    company = company_required()
    if not company: return redirect(url_for("companies_dashboard"))
    start, end = report_period(company); currency_code, factor = report_factor(company); movement = movements(company["id"], start, end); accounts_list = balances(company["id"], end)
    groups = {kind: [] for kind in ("Income", "Expense")}
    for account in accounts_list:
        if account["category"] in groups:
            debit, credit = movement[account["id"]]; amount = credit - debit if account["category"] == "Income" else debit - credit
            groups[account["category"]].append({**account, "period_balance": amount})
    totals = {kind: sum(a["period_balance"] for a in rows) for kind, rows in groups.items()}
    if factor != 1:
        for group in groups.values():
            for account in group: account["period_balance"] *= float(factor)
        totals = {kind: value * float(factor) for kind, value in totals.items()}
    export_rows = []
    for account in groups["Income"]:
        export_rows.append(["Revenue", account["name"], money(account["period_balance"])])
    export_rows.append(["", "TOTAL REVENUE", money(totals["Income"])])
    for account in groups["Expense"]:
        export_rows.append(["Expense", account["name"], money(account["period_balance"])])
    export_rows.append(["", "TOTAL EXPENSES", money(totals["Expense"])])
    export_rows.append(["", "NET PROFIT", money(totals["Income"]-totals["Expense"])])
    result = export_report(company, "Income Statement", ["Group", "Account", currency_code], export_rows, currency_code)
    if result: return result
    return render_template("income_statement.html", groups=groups, totals=totals, profit=totals["Income"]-totals["Expense"], start=start, end=end, base_currency=currency_code, report_rate=str(factor))


@app.route("/reports/financial-position")
def financial_position():
    company = company_required()
    if not company: return redirect(url_for("companies_dashboard"))
    _start, end = report_period(company); currency_code, factor = report_factor(company); data = balances(company["id"], end)
    groups = {kind: [item for item in data if item["category"] == kind] for kind in ("Asset", "Liability", "Equity")}; totals = {kind: sum(item["balance"] for item in rows) for kind, rows in groups.items()}
    difference = totals["Asset"] - totals["Liability"] - totals["Equity"]
    if abs(difference) > .005:
        if difference > 0:
            groups["Equity"].append({"id": None, "code": "", "name": "Diff in opening balance", "balance": difference, "is_difference": True}); totals["Equity"] += difference
        else:
            groups["Asset"].append({"id": None, "code": "", "name": "Diff in opening balance", "balance": -difference, "is_difference": True}); totals["Asset"] += -difference
    if factor != 1:
        for group in groups.values():
            for account in group: account["balance"] *= float(factor)
        totals = {kind: value * float(factor) for kind, value in totals.items()}
    export_rows = []
    for kind in ("Asset", "Liability", "Equity"):
        for account in groups[kind]:
            export_rows.append([kind, account["name"], money(account["balance"])])
        export_rows.append(["", f"TOTAL {kind.upper()}S", money(totals[kind])])
    result = export_report(company, "Financial Position", ["Group", "Account", currency_code], export_rows, currency_code)
    if result: return result
    return render_template("financial_position.html", groups=groups, totals=totals, end=end, base_currency=currency_code, report_rate=str(factor))


@app.route("/reports/general-ledger")
def general_ledger():
    company = company_required()
    if not company: return redirect(url_for("companies_dashboard"))
    start, end = report_period(company); currency_code, factor = report_factor(company); account_id = request.args.get("account_id", type=int); has_selection = "account_id" in request.args and account_id is not None
    accounts = db().execute("SELECT * FROM accounts WHERE company_id=? ORDER BY code", (company["id"],)).fetchall()
    sql = """SELECT e.id entry_id,e.entry_date,e.document_type,e.document_no,e.reference,e.memo,a.code,a.name,l.description,l.currency,l.fx_rate,l.debit,l.credit,l.debit*l.fx_rate base_debit,l.credit*l.fx_rate base_credit FROM journal_lines l JOIN journal_entries e ON e.id=l.entry_id JOIN accounts a ON a.id=l.account_id WHERE e.company_id=? AND e.entry_date BETWEEN ? AND ?"""; params = [company["id"], start, end]
    if has_selection and account_id:
        sql += " AND a.id=?"; params.append(account_id)
    rows = db().execute(sql + " ORDER BY e.entry_date,e.id,l.id", params).fetchall() if has_selection else []
    if factor != 1:
        rows = [{**dict(row), "base_debit": row["base_debit"] * float(factor), "base_credit": row["base_credit"] * float(factor)} for row in rows]
    export_rows = [[display_date(r["entry_date"]), r["document_type"], r["document_no"], r["name"], r["reference"] or "", r["description"] or r["memo"] or "", r["currency"], money(r["base_debit"]), money(r["base_credit"])] for r in rows]
    export_rows.append(["TOTAL", "", "", "", "", "", "", money(sum(r["base_debit"] for r in rows)), money(sum(r["base_credit"] for r in rows))])
    result = export_report(company, "General Ledger", ["Date", "Document Type", "Document No.", "Account", "Reference", "Description", "Currency", "Debit", "Credit"], export_rows, currency_code) if has_selection else None
    if result: return result
    if request.args.get("export") == "csv":
        output = StringIO(); writer = csv.writer(output); writer.writerow(["Date", "Document Type", "Document No.", "Account", "Reference", "Memo", "Description", "Currency", "Debit", "Credit", f"Debit {company['base_currency']}", f"Credit {company['base_currency']}"])
        writer.writerows([[display_date(r["entry_date"]), r["document_type"], r["document_no"], r["name"], r["reference"], r["memo"], r["description"], r["currency"], r["debit"], r["credit"], r["base_debit"], r["base_credit"]] for r in rows])
        return Response(output.getvalue(), mimetype="text/csv", headers={"Content-Disposition": "attachment; filename=general-ledger.csv"})
    return render_template("general_ledger.html", rows=rows, accounts=accounts, start=start, end=end, account_id=account_id, has_selection=has_selection, base_currency=currency_code, report_rate=str(factor))


@app.route("/reports/cash-flow")
def cash_flow():
    company = company_required()
    if not company: return redirect(url_for("companies_dashboard"))
    start, end = report_period(company); currency_code, factor = report_factor(company); selected = request.args.getlist("types") or ["Cash", "Bank", "Card"]
    selected = [item for item in selected if item in ("Cash", "Bank", "Card")]
    placeholders = ",".join("?" for _ in selected)
    rows = db().execute(f"""SELECT e.entry_date,a.code,a.name,a.cash_type,l.description,l.debit*l.fx_rate debit,l.credit*l.fx_rate credit FROM journal_lines l JOIN journal_entries e ON e.id=l.entry_id JOIN accounts a ON a.id=l.account_id WHERE e.company_id=? AND e.entry_date BETWEEN ? AND ? AND a.is_cash=1 AND a.cash_type IN ({placeholders}) ORDER BY e.entry_date""", [company["id"], start, end, *selected]).fetchall()
    total_debit, total_credit = sum(r["debit"] for r in rows), sum(r["credit"] for r in rows)
    if factor != 1:
        rows = [{**dict(row), "debit": row["debit"] * float(factor), "credit": row["credit"] * float(factor)} for row in rows]
        total_debit *= float(factor); total_credit *= float(factor)
    export_rows = [[display_date(r["entry_date"]), r["cash_type"], r["name"], r["description"] or "", money(r["debit"]), money(r["credit"])] for r in rows]
    export_rows.append(["TOTAL", "", "", "", money(total_debit), money(total_credit)])
    result = export_report(company, "Cash Flow", ["Date", "Type", "Account", "Description", "Inflow", "Outflow"], export_rows, currency_code)
    if result: return result
    return render_template("cash_flow.html", rows=rows, start=start, end=end, selected=selected, total_debit=total_debit, total_credit=total_credit, base_currency=currency_code, report_rate=str(factor))


if __name__ == "__main__":
    init_db()
    app.run(host="0.0.0.0", port=5000, debug=True)
